"""Phase 20.3 bundle — Bulk Product Importer.

Endpoints:
  POST /api/products/bulk-import/preview  — multipart upload, dry-run preview (admin)
  POST /api/products/bulk-import/commit   — commit upload (admin), revocable batch
  GET  /api/products/bulk-import/template — CSV template download
"""
from __future__ import annotations

import csv
import io
import logging
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook

from core.auth import get_current_user
from core.database import db
from services import import_batch_service as ibs
from services.audit_service import log_action

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/products-bulk-import", tags=["products"])

ADMIN_ROLES = {"admin", "admin_owner", "super_admin"}

REQUIRED_COLUMNS = ["name", "country", "category"]
OPTIONAL_COLUMNS = [
    "description", "visa_type", "visa_subclass", "assessing_body_code",
    "base_fee", "is_pre_assessment", "pre_assessment_fee_inr",
    "workflow_id", "status",
]

VALID_CATEGORIES = {
    "skilled_migration", "pr", "work", "study", "tourist", "visitor",
    "investment", "business", "dependent", "parent", "child",
    "exam_voucher", "coaching", "service_addon", "uncategorized",
}


def _is_admin(user: Dict[str, Any]) -> bool:
    role = user.get("rbac_role") or user.get("role")
    return role in ADMIN_ROLES or "*" in (user.get("permissions") or [])


def _parse_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v or "").strip().lower()
    return s in ("true", "yes", "y", "1", "t")


def _parse_int(v: Any) -> Optional[int]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _parse_rows(content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Parse CSV or XLSX into list of row dicts (lowercase column keys)."""
    fname = (filename or "").lower()
    if fname.endswith(".csv") or fname.endswith(".tsv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [{(k or "").strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader]
    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, [])
        cols = [(str(c).strip().lower() if c else "") for c in header]
        out = []
        for raw in rows_iter:
            row = {}
            for i, c in enumerate(cols):
                if c:
                    val = raw[i] if i < len(raw) else None
                    row[c] = str(val).strip() if val is not None else ""
            if any(row.values()):
                out.append(row)
        wb.close()
        return out
    raise HTTPException(status_code=400, detail=f"Unsupported file type: {filename}")


async def _validate_row(row: Dict[str, Any], idx: int) -> Dict[str, Any]:
    """Validate + normalize a single row. Returns {valid, errors, normalized}."""
    errors = []
    # Required fields
    for col in REQUIRED_COLUMNS:
        if not row.get(col):
            errors.append(f"Missing required field: {col}")
    # Category enum check
    cat = (row.get("category") or "").strip().lower()
    if cat and cat not in VALID_CATEGORIES:
        errors.append(f"Invalid category '{cat}' (must be one of {sorted(VALID_CATEGORIES)})")
    # Country format
    country = (row.get("country") or "").strip()
    # FK: workflow_id
    wf_id = (row.get("workflow_id") or "").strip()
    if wf_id:
        wf = await db["ai_workflow_templates"].find_one({"id": wf_id, "verified": True})
        if not wf:
            errors.append(f"Unknown or unverified workflow_id: {wf_id}")
    # FK: assessing_body_code (only valid for AU/NZ)
    body = (row.get("assessing_body_code") or "").strip().upper()
    if body:
        if country.upper()[:2] not in ("AU", "NZ"):
            errors.append(f"assessing_body_code valid only for AU/NZ, got country={country}")
        else:
            aa = await db["assessing_authorities"].find_one({"code": body})
            if not aa:
                errors.append(f"Unknown assessing_body_code: {body}")
    # Numeric checks
    base_fee = _parse_int(row.get("base_fee"))
    if row.get("base_fee") and base_fee is None:
        errors.append(f"Invalid base_fee: {row.get('base_fee')}")
    pa_fee = _parse_int(row.get("pre_assessment_fee_inr"))
    if row.get("pre_assessment_fee_inr") and pa_fee is None:
        errors.append(f"Invalid pre_assessment_fee_inr: {row.get('pre_assessment_fee_inr')}")

    normalized = {
        "name": (row.get("name") or "").strip(),
        "description": (row.get("description") or "").strip(),
        "category": cat or "uncategorized",
        "country": country,
        "visa_type": (row.get("visa_type") or "").strip(),
        "visa_subclass": (row.get("visa_subclass") or "").strip() or None,
        "assessing_body_code": body or None,
        "base_fee": base_fee or 0,
        "service_price": base_fee or 0,
        "is_pre_assessment": _parse_bool(row.get("is_pre_assessment")),
        "pre_assessment_fee_inr": pa_fee,
        "pre_assessment_fee_currency": "INR",
        "workflow_id": wf_id or None,
        "status": (row.get("status") or "active").strip(),
    }
    return {"row_index": idx, "valid": not errors, "errors": errors, "normalized": normalized}


async def _check_existing(name: str, country: str) -> Optional[Dict[str, Any]]:
    return await db["products"].find_one({"name": name, "country": country}, {"_id": 0})


@router.post("/preview")
async def bulk_import_preview(
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin only")
    content = await file.read()
    rows = _parse_rows(content, file.filename or "upload.csv")
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file or no rows detected")

    out = []
    new_count = update_count = invalid_count = 0
    for i, row in enumerate(rows, 1):
        v = await _validate_row(row, i)
        existing = None
        if v["valid"]:
            existing = await _check_existing(v["normalized"]["name"], v["normalized"]["country"])
        action = "invalid" if not v["valid"] else ("update" if existing else "new")
        if action == "new": new_count += 1
        elif action == "update": update_count += 1
        else: invalid_count += 1
        out.append({
            **v, "action": action,
            "existing_id": existing.get("id") if existing else None,
        })
    return {
        "total_rows": len(rows),
        "new": new_count, "update": update_count, "invalid": invalid_count,
        "rows": out,
        "preview_id": str(uuid.uuid4()),
        "filename": file.filename,
    }


@router.post("/commit")
async def bulk_import_commit(
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin only")
    content = await file.read()
    rows = _parse_rows(content, file.filename or "upload.csv")
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    user_id = str(current_user.get("id") or "admin")
    user_name = str(current_user.get("name") or current_user.get("email") or "admin")
    batch = await ibs.open_batch(
        db, ingestion_path="phase_20.3_products.bulk_import",
        endpoint="POST /api/products/bulk-import/commit",
        uploaded_by=user_id, uploaded_by_name=user_name,
        file_name=file.filename, file_hash=ibs.file_sha256(content),
        file_size_bytes=len(content), target_collection="products",
    )

    created = updated = skipped = 0
    errors = []
    now = datetime.now(timezone.utc)

    for i, row in enumerate(rows, 1):
        v = await _validate_row(row, i)
        if not v["valid"]:
            skipped += 1
            errors.append({"row": i, "errors": v["errors"]})
            continue
        normalized = v["normalized"]
        existing = await _check_existing(normalized["name"], normalized["country"])
        if existing:
            await db["products"].update_one(
                {"id": existing["id"]},
                {"$set": {**normalized, "updated_at": now}},
            )
            ibs.record_update(batch, existing["id"], normalized,
                              {k: existing.get(k) for k in normalized.keys()})
            updated += 1
        else:
            new_id = str(uuid.uuid4())
            new_doc = {**normalized, "id": new_id, "created_at": now, "updated_at": now,
                       "archived_at": None, "workflow_steps_count": 0,
                       "commissions_v2": None, "commission_type": "percentage",
                       "commission_rate": 0.0, "commission_tiers": [],
                       "cost_allocations": [], "success_bonuses": [],
                       "computed": None}
            await db["products"].insert_one(new_doc)
            ibs.record_create(batch, new_id, new_doc)
            created += 1

    await ibs.close_batch(db, batch, total_rows=len(rows), status="committed")
    await log_action(db, action="products.bulk_import",
                     user_id=user_id, user_name=user_name, severity="info",
                     summary={"batch_id": batch["batch_id"], "created": created,
                              "updated": updated, "skipped": skipped, "total": len(rows)})
    return {
        "ok": True, "batch_id": batch["batch_id"],
        "total_rows": len(rows), "created": created, "updated": updated,
        "skipped": skipped, "errors": errors[:20],  # cap output
    }


@router.get("/template")
async def download_template(current_user: Dict[str, Any] = Depends(get_current_user)):
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin only")
    header = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
    sample = [
        ["Australia PR Skilled Migration", "Australia", "pr",
         "Australia 189/190/491 skilled migration pathway", "Skilled Migration",
         "189", "ACS", "150000", "false", "", "australia_pr_verified", "active"],
        ["Canada PR Express Entry", "Canada", "pr",
         "Canada Express Entry pathway", "PR", "EE", "", "180000",
         "false", "", "canada_pr_verified", "active"],
        ["UK Student Visa", "UK", "study",
         "UK student visa for international students", "Study", "T4", "",
         "75000", "true", "3000", "", "active"],
    ]
    csv_text = io.StringIO()
    writer = csv.writer(csv_text)
    writer.writerow(header)
    for s in sample:
        # Pad sample to match header length
        writer.writerow(s + [""] * (len(header) - len(s)))
    from fastapi.responses import Response
    return Response(
        content=csv_text.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="products_bulk_template.csv"'},
    )
