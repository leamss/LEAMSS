"""Phase 19.4 — Universal Data Import Hub.

Endpoints:
  POST   /api/data-import/upload                  — multipart file → returns file_id + auto-detected type
  POST   /api/data-import/{file_id}/parse-preview — dry-run, returns row count + sample
  POST   /api/data-import/{file_id}/commit        — idempotent commit
  GET    /api/data-import/history                 — paginated list
  DELETE /api/data-import/{file_id}               — admin only, cascade delete

Auth: admin only on all routes (audit-logged on commit).
"""
from __future__ import annotations

import logging
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from motor.motor_asyncio import AsyncIOMotorDatabase
from openpyxl import load_workbook

from core.auth import get_current_user
from scrapers.base import _db as db
from parsers.jsa import (
    occupation_profiles as p_occ_profiles,
    employment_projections as p_emp_proj,
    sa4_ratings as p_sa4,
    industry_data as p_industry,
    vacancy_report as p_vacancy,
)
from services import jsa_importer

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/data-import", tags=["data-import"])

STORAGE_DIR = Path(os.environ.get("DATA_IMPORT_STORAGE", "/app/backend/storage/jsa_imports"))
STORAGE_DIR.mkdir(parents=True, exist_ok=True)

PARSER_REGISTRY = {
    "occupation_profiles": p_occ_profiles,
    "employment_projections": p_emp_proj,
    "sa4_ratings": p_sa4,
    "industry_data": p_industry,
    "vacancy_report": p_vacancy,
}

COMMITTER_REGISTRY = {
    "occupation_profiles": jsa_importer.commit_occupation_profiles,
    "employment_projections": jsa_importer.commit_employment_projections,
    "sa4_ratings": jsa_importer.commit_sa4_ratings,
    "industry_data": jsa_importer.commit_industry_data,
    "vacancy_report": jsa_importer.commit_vacancy_report,
}


def _is_admin(user: Dict[str, Any]) -> bool:
    role = (user or {}).get("role", "")
    return role in ("admin", "super_admin")


def _require_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Admin role required")
    return user


def _detect_type(path: Path) -> str:
    """Auto-detect file type by extension + inspect."""
    # PDF branch — currently only Vacancy Report supported
    if path.suffix.lower() == ".pdf":
        try:
            import pdfplumber  # noqa: PLC0415
            with pdfplumber.open(str(path)) as pdf:
                page1 = pdf.pages[0].extract_text() or ""
            if "vacancy report" in page1.lower() or "internet vacancy index" in page1.lower():
                return "vacancy_report"
        except Exception as e:  # noqa: BLE001
            logger.exception("pdf type detection failed: %s", e)
        return "unknown"
    # XLSX branch
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        first_sheet = wb[sheet_names[0]]
        first_rows = []
        for i, row in enumerate(first_sheet.iter_rows(values_only=True, max_row=15)):
            first_rows.append(list(row))
        wb.close()
        return jsa_importer.detect_file_type(sheet_names, first_rows)
    except Exception as e:  # noqa: BLE001
        logger.exception("type detection failed: %s", e)
        return "unknown"


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(_require_admin),
) -> Dict[str, Any]:
    """Save uploaded file + auto-detect its type. Returns file_id."""
    file_id = str(uuid.uuid4())
    safe_name = (file.filename or "upload.xlsx").replace("/", "_")[:255]
    storage_path = STORAGE_DIR / f"{file_id}__{safe_name}"

    # Stream-write to disk
    content = await file.read()
    storage_path.write_bytes(content)
    file_size = storage_path.stat().st_size

    detected_type = _detect_type(storage_path) if safe_name.lower().endswith((".xlsx", ".pdf")) else "unsupported_extension"

    doc = {
        "id": file_id,
        "filename": safe_name,
        "stored_path": str(storage_path),
        "size_bytes": file_size,
        "detected_type": detected_type,
        "status": "uploaded",
        "uploaded_by": user.get("id") or user.get("email"),
        "uploaded_at": datetime.now(timezone.utc),
    }
    await db["import_files"].insert_one(doc)

    return {
        "file_id": file_id,
        "filename": safe_name,
        "size_bytes": file_size,
        "detected_type": detected_type,
        "supported": detected_type in PARSER_REGISTRY,
    }


async def _get_file(file_id: str) -> Dict[str, Any]:
    doc = await db["import_files"].find_one({"id": file_id})
    if not doc:
        raise HTTPException(status_code=404, detail="file not found")
    return doc


@router.post("/{file_id}/parse-preview")
async def parse_preview(
    file_id: str, user: Dict[str, Any] = Depends(_require_admin)
) -> Dict[str, Any]:
    """Dry-run parse — returns sample 5 rows + total count, no DB writes."""
    doc = await _get_file(file_id)
    ftype = doc["detected_type"]
    if ftype not in PARSER_REGISTRY:
        raise HTTPException(status_code=400, detail=f"unsupported file type: {ftype}")
    parser = PARSER_REGISTRY[ftype]
    try:
        summary = parser.parse_summary(doc["stored_path"])
    except Exception as e:  # noqa: BLE001
        logger.exception("parse-preview failed")
        raise HTTPException(status_code=500, detail=f"parse failed: {e}") from e

    return {
        "file_id": file_id,
        "filename": doc["filename"],
        "detected_type": ftype,
        "row_count": summary["row_count"],
        "source": summary["source"],
        "source_url": summary["source_url"],
        "sample": summary["sample"],
        "honest_note": summary.get("honest_note"),
    }


@router.post("/{file_id}/commit")
async def commit_file(
    file_id: str, user: Dict[str, Any] = Depends(_require_admin)
) -> Dict[str, Any]:
    """Idempotent commit — runs parser + writes to DB. Always safe to re-run.
    Phase 19.6 — registers an `import_batches` row (audit-only; granular revoke
    is not available for JSA bulk upserts since pre-state is not captured here)."""
    doc = await _get_file(file_id)
    ftype = doc["detected_type"]
    if ftype not in PARSER_REGISTRY:
        raise HTTPException(status_code=400, detail=f"unsupported file type: {ftype}")
    parser = PARSER_REGISTRY[ftype]
    committer = COMMITTER_REGISTRY[ftype]

    # Phase 19.6 — open batch BEFORE running committer (audit-only)
    from services import import_batch_service as ibs
    try:
        file_bytes = Path(doc["stored_path"]).read_bytes()
        file_hash = ibs.file_sha256(file_bytes)
        file_size = len(file_bytes)
    except Exception:  # noqa: BLE001
        file_hash, file_size = "", 0
    _TARGET_BY_TYPE = {
        "occupation_profiles": "occupation_master",
        "employment_projections": "occupation_master",
        "sa4_ratings": "regional_labour_market",
        "industry_data": "industry_master",
        "vacancy_report": "vacancy_snapshots",
    }
    batch = await ibs.open_batch(
        db,
        ingestion_path=f"phase_19.4_data_import.{ftype}",
        endpoint=f"POST /api/data-import/{{file_id}}/commit ({ftype})",
        uploaded_by=str(user.get("id") or user.get("email") or "admin"),
        uploaded_by_name=str(user.get("name") or user.get("email") or "admin"),
        file_name=doc.get("filename", "upload"),
        file_hash=file_hash,
        file_size_bytes=file_size,
        target_collection=_TARGET_BY_TYPE.get(ftype, "unknown"),
    )

    started = datetime.now(timezone.utc)
    try:
        parsed = list(parser.parse_workbook(doc["stored_path"]))
        summary = await committer(db, parsed)
    except Exception as e:  # noqa: BLE001
        logger.exception("commit failed")
        await db["import_files"].update_one(
            {"id": file_id},
            {"$set": {"status": "failed", "last_error": str(e)[:300]}},
        )
        # Phase 19.6 — close batch as failed
        try:
            await db["import_batches"].update_one(
                {"batch_id": batch["batch_id"]},
                {"$set": {"status": "failed", "is_revocable": False,
                          "audit_only": True, "non_revocable_reason": "commit_failed"}},
            )
        except Exception:  # noqa: BLE001
            pass
        raise HTTPException(status_code=500, detail=f"commit failed: {e}") from e

    finished = datetime.now(timezone.utc)
    elapsed_s = (finished - started).total_seconds()

    await db["import_files"].update_one(
        {"id": file_id},
        {"$set": {
            "status": "committed",
            "committed_at": finished,
            "commit_summary": summary,
            "elapsed_seconds": elapsed_s,
        }},
    )

    await jsa_importer.audit_log(
        db, str(user.get("id") or user.get("email")),
        f"data_import.commit.{ftype}",
        {"file_id": file_id, "filename": doc["filename"], "batch_id": batch["batch_id"], **summary},
    )

    # Phase 19.6 — close batch as audit-only with parsed counts
    _total_rows = int(summary.get("parsed_records") or summary.get("parsed_4digit_records") or 0)
    _created = int(summary.get("regions_upserted", 0) or summary.get("industries_upserted", 0)
                    or summary.get("snapshots_upserted", 0) or 0)
    _updated = int(summary.get("regions_modified", 0) or summary.get("industries_modified", 0)
                    or summary.get("snapshots_modified", 0) or summary.get("occupations_updated", 0) or 0)
    _skipped = int(summary.get("occupations_skipped_no_4digit_match", 0) or 0)
    await ibs.close_batch(db, batch, total_rows=_total_rows, status="committed")
    await db["import_batches"].update_one(
        {"batch_id": batch["batch_id"]},
        {"$set": {
            "is_revocable": False, "audit_only": True,
            "non_revocable_reason": "bulk_upsert_audit_only",
            "counts.created": _created, "counts.updated": _updated, "counts.skipped": _skipped,
            "summary_payload": summary,
        }},
    )

    return {
        "file_id": file_id,
        "type": ftype,
        "summary": summary,
        "elapsed_seconds": elapsed_s,
        "batch_id": batch["batch_id"],
    }


@router.get("/history")
async def list_history(
    page: int = Query(1, ge=1), limit: int = Query(20, ge=1, le=100),
    user: Dict[str, Any] = Depends(_require_admin),
) -> Dict[str, Any]:
    skip = (page - 1) * limit
    total = await db["import_files"].count_documents({})
    cursor = db["import_files"].find({}, {"_id": 0}).sort("uploaded_at", -1).skip(skip).limit(limit)
    items = [d async for d in cursor]
    # Stringify datetime for JSON
    for it in items:
        for k in ("uploaded_at", "committed_at"):
            v = it.get(k)
            if isinstance(v, datetime):
                it[k] = v.isoformat()
    return {"page": page, "limit": limit, "total": total, "items": items}


@router.delete("/{file_id}")
async def delete_import(
    file_id: str, user: Dict[str, Any] = Depends(_require_admin)
) -> Dict[str, Any]:
    """Cascade delete — removes DB row + on-disk file. Does NOT rollback committed data."""
    doc = await _get_file(file_id)
    try:
        Path(doc["stored_path"]).unlink(missing_ok=True)
    except Exception as e:  # noqa: BLE001
        logger.warning("delete failed for file %s: %s", file_id, e)
    await db["import_files"].delete_one({"id": file_id})
    return {"deleted": True, "file_id": file_id}


# Phase 19.4c — Vacancy snapshot quick-read endpoints (read-only, admin)
@router.get("/vacancy/latest")
async def vacancy_latest(user: Dict[str, Any] = Depends(_require_admin)) -> Dict[str, Any]:
    """Return the most recent imported `vacancy_snapshots` doc (is_latest=True)."""
    doc = await db["vacancy_snapshots"].find_one({"is_latest": True}, {"_id": 0})
    if not doc:
        return {"snapshot": None}
    return {"snapshot": doc}


@router.get("/industries")
async def list_industries(user: Dict[str, Any] = Depends(_require_admin)) -> Dict[str, Any]:
    """Phase 19.4c — Return all 19 ANZSIC industry docs from industry_master."""
    items = []
    async for ind in db["industry_master"].find({}, {"_id": 0}).sort("employed_count", -1):
        items.append(ind)
    return {"count": len(items), "items": items}
