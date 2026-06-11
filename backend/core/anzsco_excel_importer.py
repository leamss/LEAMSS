"""Phase 7.1 — ANZSCO Feb 2026 Excel Importer.

Parses the official ABS ANZSCO Feb 2026 occupation profiles Excel workbook
into MongoDB collection `anzsco_4digit_master`.

The Excel has 9 data tables:
  Table_1: Overview (employed count, demographics, earnings, growth)
  Table_2: Occupation descriptions
  Table_3: Occupation tasks (multi-row per code)
  Table_4: Earnings & hours
  Table_5: Industries (multi-row per code, ranked)
  Table_6: States & Territories distribution
  Table_7: Age profile distribution
  Table_8: Education attainment distribution
  Table_9: All-occupation aggregates (not stored per-code)

ANZSCO uses 4-digit codes (e.g., 2613 = Software & Applications Programmers).
6-digit codes in our `occupation_master` (e.g., 261313 = Software Engineer)
inherit their 4-digit parent's profile.

Idempotent: Re-running upserts by `code`, never deletes.
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from core.database import db

logger = logging.getLogger(__name__)

ANZSCO_4DIGIT = db["anzsco_4digit_master"]

# Sheet header row is always Row 7 (per ABS Feb 2026 layout). Data starts Row 8.
HEADER_ROW = 7
DATA_START_ROW = 8

# Phase 17.0.2 — schema contract exposed for pre-validation by the upload router.
# Table_9 is "All-occupation aggregates" — not strictly required for the per-code
# parsers to succeed, so we leave it out of the must-have list.
REQUIRED_SHEETS: Tuple[str, ...] = (
    "Table_1", "Table_2", "Table_3", "Table_4", "Table_5", "Table_6", "Table_7", "Table_8",
)
# Header row 7 in Table_1 MUST contain (case-insensitive, trimmed) at least
# one match for each role. Accepts ABS variants seen in past releases.
REQUIRED_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "code":  ("code", "anzsco code", "anzsco", "occupation code"),
    "title": ("title", "occupation", "occupation title", "name"),
}


def _clean(v: Any) -> Any:
    """Normalize cell values — 'N/A' → None, strip strings, leave numbers."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s.upper() in ("N/A", "NA", "-", ""):
            return None
        return s
    return v


def _to_int(v: Any) -> Optional[int]:
    v = _clean(v)
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _to_float(v: Any) -> Optional[float]:
    v = _clean(v)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_str(v: Any) -> Optional[str]:
    v = _clean(v)
    return None if v is None else str(v)


def parse_table_1_overview(ws) -> Dict[str, Dict[str, Any]]:
    """Returns {code: {employed, part_time_share_pct, female_share_pct,
    median_weekly_earnings_aud, median_age, annual_employment_growth}}.
    """
    out: Dict[str, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        if not code or not code.isdigit():
            continue
        title = _to_str(row[1])
        if not title:
            continue
        out[code] = {
            "code": code,
            "title": title,
            "employed_count": _to_int(row[2]),
            "part_time_share_pct": _to_float(row[3]),
            "female_share_pct": _to_float(row[4]),
            "median_weekly_earnings_aud": _to_int(row[5]),
            "median_age": _to_int(row[6]),
            "annual_employment_growth": _to_int(row[7]),
        }
    return out


def parse_table_2_descriptions(ws) -> Dict[str, str]:
    """Returns {code: description}."""
    out: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        desc = _to_str(row[2])
        if code and code.isdigit() and desc:
            out[code] = desc
    return out


def parse_table_3_tasks(ws) -> Dict[str, List[str]]:
    """Returns {code: [task1, task2, ...]}. Multi-row per code."""
    out: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        task = _to_str(row[2])
        if not (code and code.isdigit() and task):
            continue
        out.setdefault(code, []).append(task)
    return out


def parse_table_4_earnings(ws) -> Dict[str, Dict[str, Any]]:
    """Returns {code: {full_time_share_pct, avg_full_time_hours_per_week,
    median_full_time_weekly_aud, median_full_time_hourly_aud}}.
    """
    out: Dict[str, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        if not code or not code.isdigit():
            continue
        out[code] = {
            "full_time_share_pct": _to_float(row[2]),
            "avg_full_time_hours_per_week": _to_float(row[3]),
            "median_full_time_weekly_aud": _to_int(row[4]),
            "median_full_time_hourly_aud": _to_int(row[5]),
        }
    return out


def parse_table_5_industries(ws) -> Dict[str, List[str]]:
    """Returns {code: [industry1, industry2, ...]}. Order = rank."""
    out: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        industry = _to_str(row[2])
        if not (code and code.isdigit() and industry):
            continue
        out.setdefault(code, []).append(industry)
    return out


def parse_table_6_states(ws) -> Dict[str, Dict[str, float]]:
    """Returns {code: {NSW, VIC, QLD, SA, WA, TAS, ACT, NT}}."""
    state_keys = ["NSW", "VIC", "QLD", "SA", "WA", "TAS", "ACT", "NT"]
    out: Dict[str, Dict[str, float]] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        if not code or not code.isdigit():
            continue
        out[code] = {k: _to_float(row[2 + i]) for i, k in enumerate(state_keys) if (2 + i) < len(row)}
    return out


def parse_table_7_age(ws) -> Dict[str, Dict[str, float]]:
    """Returns {code: {age_15_19, age_20_24, age_25_34, age_35_44, age_45_54, age_55_59, age_60_64, age_65_plus}}."""
    bins = ["age_15_19", "age_20_24", "age_25_34", "age_35_44", "age_45_54", "age_55_59", "age_60_64", "age_65_plus"]
    out: Dict[str, Dict[str, float]] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        if not code or not code.isdigit():
            continue
        out[code] = {k: _to_float(row[2 + i]) for i, k in enumerate(bins) if (2 + i) < len(row)}
    return out


def parse_table_8_education(ws) -> Dict[str, Dict[str, float]]:
    """Returns {code: {post_grad, bachelor, diploma, cert_3_4, year_12, year_11, year_10_or_below}}."""
    bins = ["post_grad", "bachelor", "diploma", "cert_3_4", "year_12", "year_11", "year_10_or_below"]
    out: Dict[str, Dict[str, float]] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        code = _to_str(row[0])
        if not code or not code.isdigit():
            continue
        out[code] = {k: _to_float(row[2 + i]) for i, k in enumerate(bins) if (2 + i) < len(row)}
    return out


async def import_anzsco_excel(file_path: str, imported_by: str = "admin") -> Dict[str, Any]:
    """Main entry — parse Excel and upsert into anzsco_4digit_master.

    Args:
        file_path: Absolute path to the .xlsx file.
        imported_by: User ID for audit.

    Returns:
        {imported, updated, skipped, errors, duration_seconds}
    """
    start = datetime.now(timezone.utc)
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(str(p), data_only=True, read_only=True)
    sheets = {sn for sn in wb.sheetnames}
    for required in ("Table_1", "Table_2", "Table_3", "Table_4", "Table_5", "Table_6", "Table_7", "Table_8"):
        if required not in sheets:
            raise ValueError(f"Required sheet '{required}' missing in workbook")

    logger.info("Parsing Excel tables…")
    overview = parse_table_1_overview(wb["Table_1"])
    descriptions = parse_table_2_descriptions(wb["Table_2"])
    tasks = parse_table_3_tasks(wb["Table_3"])
    earnings = parse_table_4_earnings(wb["Table_4"])
    industries = parse_table_5_industries(wb["Table_5"])
    states = parse_table_6_states(wb["Table_6"])
    age = parse_table_7_age(wb["Table_7"])
    education = parse_table_8_education(wb["Table_8"])
    logger.info("Parsed %d overview rows · %d task rows-merged · %d industry rows-merged",
                len(overview), sum(len(v) for v in tasks.values()), sum(len(v) for v in industries.values()))

    imported = 0
    updated = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []

    for code, ov in overview.items():
        try:
            doc = {
                "code": code,
                "title": ov["title"],
                "classification_system": "ANZSCO",
                "level": "4_digit",
                "description": descriptions.get(code),
                "tasks": tasks.get(code, []),
                "anzsco_profile": {
                    "employed_count": ov.get("employed_count"),
                    "part_time_share_pct": ov.get("part_time_share_pct"),
                    "female_share_pct": ov.get("female_share_pct"),
                    "median_weekly_earnings_aud": ov.get("median_weekly_earnings_aud"),
                    "median_age": ov.get("median_age"),
                    "annual_employment_growth": ov.get("annual_employment_growth"),
                    **(earnings.get(code, {})),
                },
                "industries_ranked": industries.get(code, []),
                "state_distribution": states.get(code, {}),
                "age_profile": age.get(code, {}),
                "education_distribution": education.get(code, {}),
                "data_source": {
                    "label": "ABS ANZSCO Feb 2026 Occupation Profiles",
                    "url": "https://www.jobsandskills.gov.au/data/occupation-and-industry-profiles/occupations",
                    "reference_period": "February 2026",
                    "imported_at": datetime.now(timezone.utc),
                    "imported_by": imported_by,
                },
                "status": "active",
                "updated_at": datetime.now(timezone.utc),
            }

            existing = await ANZSCO_4DIGIT.find_one({"code": code}, {"_id": 1})
            if existing:
                await ANZSCO_4DIGIT.update_one({"code": code}, {"$set": doc})
                updated += 1
            else:
                doc["created_at"] = datetime.now(timezone.utc)
                await ANZSCO_4DIGIT.insert_one(doc)
                imported += 1
        except Exception as e:
            logger.exception("Row error for code %s", code)
            errors.append({"code": code, "error": str(e)})
            skipped += 1

    duration = (datetime.now(timezone.utc) - start).total_seconds()
    summary = {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:50],  # cap to first 50
        "total_processed": len(overview),
        "duration_seconds": round(duration, 2),
        "source": "ABS ANZSCO Feb 2026",
    }
    logger.info("Excel import complete: %s", summary)
    return summary


async def get_4digit_parent_code(six_digit_code: str) -> Optional[str]:
    """Resolves a 6-digit ANZSCO code to its 4-digit parent."""
    if not six_digit_code or len(six_digit_code) < 4:
        return None
    return six_digit_code[:4]
