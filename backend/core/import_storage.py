"""Phase 17.0 — Persistent storage for admin-uploaded import artefacts.

Replaces the legacy `/tmp/anzsco_feb2026.xlsx` hardcoded path which got wiped
on every container restart, leaving the "Re-import Excel" button broken.

Public surface (callable from routers):
  • STORAGE_ROOT                          — absolute path under /app/backend/storage/imports
  • save_import_file(bytes, …)            — write to disk + create/update import_files row
  • get_latest_file(source_type)          — fetch most-recent {is_latest: True} doc
  • list_files(source_type, limit)        — paginated history (newest first)
  • update_last_import_summary(...)       — record import outcome on the file row
  • PUBLIC_FIELDS                         — whitelist for client responses (NO storage_path)

Design notes:
  • Files live at /app/backend/storage/imports/{source_type}/
    Filename pattern: {source_type}_{YYYYMMDD_HHMMSS}_{sanitised_original}.xlsx
  • Permissions: 600 (owner read+write only) — admin-only artefacts.
  • Hash-based dedupe: if a new upload's sha256 matches the existing latest,
    we reuse that file_id (no duplicate disk write, no new row).
  • Retention: keep last 10 per source_type, prune older from disk AND DB.
  • Response sanitisation: PUBLIC_FIELDS strips `storage_path` from any payload
    that crosses the API boundary — server-internal paths NEVER reach the client.
"""
from __future__ import annotations

import asyncio
import hashlib
import logging
import os
import re
import stat
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# ─── Constants ──────────────────────────────────────────────────────────────
STORAGE_ROOT = Path(__file__).resolve().parents[1] / "storage" / "imports"
DEFAULT_RETENTION = 10  # keep newest N files per source_type

# Whitelist of fields safe to return to API clients.
# `storage_path` is deliberately excluded — it's a server-internal absolute path.
PUBLIC_FIELDS = {
    "id",
    "source_type",
    "filename_original",
    "filename_stored",
    "size_bytes",
    "sha256",
    "uploaded_by",
    "uploaded_by_name",
    "uploaded_at",
    "is_latest",
    "status",
    "last_import_summary",
    "last_imported_at",
}


# ─── Init ───────────────────────────────────────────────────────────────────
def ensure_storage_dirs() -> None:
    """Idempotent — called on backend startup. Creates the imports root."""
    STORAGE_ROOT.mkdir(parents=True, exist_ok=True)
    try:
        # Tighten dir permissions to owner-only (700) — admin-only artefacts
        os.chmod(STORAGE_ROOT, stat.S_IRWXU)
    except OSError as e:
        logger.warning("Could not chmod storage root: %s", e)


# ─── Utilities ──────────────────────────────────────────────────────────────
def _sanitise_filename(name: str) -> str:
    """Strip path separators + any non [A-Za-z0-9._-]; cap length to 80."""
    base = os.path.basename(name or "upload.xlsx")
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._-") or "upload.xlsx"
    return base[:80]


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def public_view(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Return a client-safe projection of an import_files document.

    Strips `_id`, `storage_path`, and any other server-internal field. Returns
    None if input is None so callers can chain ``return public_view(doc)``.
    """
    if not doc:
        return None
    out: Dict[str, Any] = {}
    for k in PUBLIC_FIELDS:
        if k not in doc:
            continue
        v = doc[k]
        if isinstance(v, datetime):
            v = v.isoformat()
        out[k] = v
    return out


def _hash_bytes(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()


# ─── Core ops ───────────────────────────────────────────────────────────────
async def save_import_file(
    *,
    db,
    source_type: str,
    data: bytes,
    filename_original: str,
    uploaded_by: str,
    uploaded_by_name: str,
    retention: int = DEFAULT_RETENTION,
) -> Dict[str, Any]:
    """Persist uploaded bytes + create/refresh `import_files` row.

    Returns the *stored* doc (with internal fields). Callers MUST run it through
    :func:`public_view` before returning to clients.

    Dedupe semantics: if the new bytes hash equals the existing latest row's
    hash for the same source_type, no new file is written — we reuse that row.
    """
    ensure_storage_dirs()
    coll = db["import_files"]
    sha = _hash_bytes(data)

    # Dedupe-vs-latest
    existing_latest = await coll.find_one(
        {"source_type": source_type, "is_latest": True},
        {"_id": 0},
    )
    if existing_latest and existing_latest.get("sha256") == sha:
        logger.info(
            "[import_storage] Dedupe hit for %s — same sha256, reusing file_id=%s",
            source_type,
            existing_latest["id"],
        )
        return existing_latest

    # Demote any prior is_latest rows for this source_type.
    await coll.update_many(
        {"source_type": source_type, "is_latest": True},
        {"$set": {"is_latest": False}},
    )

    # Compose storage path.
    src_dir = STORAGE_ROOT / source_type
    src_dir.mkdir(parents=True, exist_ok=True)
    try:
        os.chmod(src_dir, stat.S_IRWXU)
    except OSError:
        pass

    safe_name = _sanitise_filename(filename_original)
    stored_name = f"{source_type}_{_timestamp()}_{safe_name}"
    storage_path = src_dir / stored_name

    # Write bytes atomically (temp → rename) so a partial write can't leak.
    tmp_path = storage_path.with_suffix(storage_path.suffix + ".part")
    with open(tmp_path, "wb") as f:
        f.write(data)
    os.replace(tmp_path, storage_path)
    try:
        os.chmod(storage_path, stat.S_IRUSR | stat.S_IWUSR)
    except OSError:
        pass

    now = datetime.now(timezone.utc).isoformat()
    doc: Dict[str, Any] = {
        "id": str(uuid.uuid4()),
        "source_type": source_type,
        "filename_original": safe_name,
        "filename_stored": stored_name,
        "storage_path": str(storage_path),  # server-internal — never sent to client
        "size_bytes": len(data),
        "sha256": sha,
        "uploaded_by": uploaded_by,
        "uploaded_by_name": uploaded_by_name,
        "uploaded_at": now,
        "is_latest": True,
        "status": "ready",
        "last_import_summary": None,
        "last_imported_at": None,
    }
    await coll.insert_one(doc)

    # Retention prune (keep newest N per source_type)
    await _prune_old(coll, source_type, retention)
    return doc


async def _prune_old(coll, source_type: str, keep: int) -> None:
    """Delete on-disk files + DB rows beyond the retention window."""
    cursor = coll.find(
        {"source_type": source_type},
        {"_id": 0, "id": 1, "storage_path": 1, "uploaded_at": 1, "is_latest": 1},
    ).sort("uploaded_at", -1)
    rows: List[Dict[str, Any]] = [d async for d in cursor]
    to_delete = rows[keep:]
    for r in to_delete:
        sp = r.get("storage_path")
        if sp:
            try:
                os.unlink(sp)
            except OSError as e:
                logger.warning("[import_storage] Could not unlink %s: %s", sp, e)
        await coll.delete_one({"id": r["id"]})
    if to_delete:
        logger.info(
            "[import_storage] Pruned %d old %s files (retention=%d)",
            len(to_delete),
            source_type,
            keep,
        )


async def get_latest_file(db, source_type: str) -> Optional[Dict[str, Any]]:
    """Return the full doc (server-internal) for the latest file of a source_type,
    or None. Callers must run through :func:`public_view` before responding."""
    return await db["import_files"].find_one(
        {"source_type": source_type, "is_latest": True},
        {"_id": 0},
    )


async def list_files(db, source_type: Optional[str], limit: int = 20) -> List[Dict[str, Any]]:
    """History list (newest first), client-safe via public_view."""
    q: Dict[str, Any] = {}
    if source_type:
        q["source_type"] = source_type
    rows: List[Dict[str, Any]] = []
    cursor = db["import_files"].find(q, {"_id": 0}).sort("uploaded_at", -1).limit(int(limit))
    async for d in cursor:
        rows.append(public_view(d) or {})
    return rows


async def update_last_import_summary(
    *,
    db,
    file_id: str,
    summary: Dict[str, Any],
    status: str = "imported",
) -> None:
    """Patch the file row with the outcome of an import run."""
    await db["import_files"].update_one(
        {"id": file_id},
        {
            "$set": {
                "last_import_summary": summary,
                "last_imported_at": datetime.now(timezone.utc).isoformat(),
                "status": status,
            }
        },
    )


async def ensure_indexes(db) -> None:
    """Create indexes if missing. Idempotent."""
    coll = db["import_files"]
    await coll.create_index(
        [("source_type", 1), ("is_latest", 1)],
        name="source_type_is_latest",
    )
    await coll.create_index(
        [("uploaded_at", -1)],
        name="uploaded_at_desc",
    )
    await coll.create_index([("id", 1)], unique=True, name="id_unique")


# ─── Phase 17.0.1 — Upload error classification ──────────────────────────────
# Shared helper used by all import endpoints to convert parser exceptions into
# user-friendly 4xx responses (so client-side junk uploads don't surface as 500).


class InvalidExcelSchemaError(ValueError):
    """Raised by :func:`validate_xlsx_schema` for client-fault schema problems
    (wrong sheets / missing required columns / no data rows). Mapped to HTTP
    400 by :func:`classify_upload_error`."""


def classify_upload_error(exc: BaseException) -> tuple[int, str]:
    """Decide whether an exception thrown during file parse/import is a CLIENT
    error (bad upload → 400) or a SERVER error (true bug → 500). Returns
    ``(status_code, sanitised_message)``. Never leaks paths or stack traces."""
    name = type(exc).__name__
    msg = str(exc)
    # Phase 17.0.2 — explicit schema-error subclass always maps to 400.
    if isinstance(exc, InvalidExcelSchemaError):
        return (400, msg)
    # zipfile.BadZipFile / openpyxl.InvalidFileException — file is not a real xlsx
    if name in ("BadZipFile", "InvalidFileException", "ParserError"):
        return (400, "Uploaded file is not a valid .xlsx workbook. Please upload a real Excel file.")
    # ValueError raised by the importer when a required sheet is missing.
    if name == "ValueError" and "Required sheet" in msg and "missing" in msg:
        return (
            400,
            "Excel file is missing one of the required sheets (Table_1 to Table_8). "
            "Please upload a valid ANZSCO workbook (ABS Feb 2026 format).",
        )
    # KeyError raised when load_workbook can't find an expected sheet name.
    if name == "KeyError" and "Worksheet" in msg:
        return (400, "Uploaded file is missing required ANZSCO sheets (Table_1, Table_2, ...). Please upload the official ABS workbook.")
    # Phase 17.0.2 — per-row column shape errors inside the importer
    # (`row[2]` on a 2-column row → IndexError). These are user-fault uploads
    # with the wrong column layout, not server bugs.
    if name == "IndexError" and "tuple index" in msg:
        return (
            400,
            "Excel file has the wrong column layout in row 7. "
            "Please upload a valid ANZSCO workbook (ABS Feb 2026 format).",
        )
    if name == "EmptyFileError" or "empty" in msg.lower() and "file" in msg.lower():
        return (400, "Uploaded file appears to be empty.")
    # Anything else → genuine server fault
    return (500, f"Excel import failed: {msg}")


def validate_xlsx_bytes(data: bytes, required_sheets: Optional[List[str]] = None) -> None:
    """Cheap in-memory validation BEFORE touching disk. Raises ValueError with
    a user-friendly message if the bytes are not a valid xlsx (and bad files
    therefore never get persisted to storage)."""
    import io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        if required_sheets:
            present = set(wb.sheetnames)
            missing = [s for s in required_sheets if s not in present]
            if missing:
                raise ValueError(
                    f"Uploaded file is missing required sheet(s): {', '.join(missing)}. "
                    "Please upload the official ABS ANZSCO workbook."
                )
    except ValueError:
        raise
    except Exception as e:  # noqa: BLE001 — convert all parse failures uniformly
        code, msg = classify_upload_error(e)
        # We only call this from upload paths, so 500-class faults here mean the
        # bytes truly can't be opened — still a client problem at this stage.
        raise ValueError(msg if code == 400 else "Uploaded file could not be parsed as a valid .xlsx workbook.") from e


def validate_xlsx_schema(
    data: bytes,
    *,
    required_sheets: List[str],
    header_row: int,
    primary_sheet: str,
    required_header_aliases: Dict[str, Any],
) -> None:
    """Phase 17.0.2 — schema-shape pre-check. Raises :class:`InvalidExcelSchemaError`
    (a client-fault subclass of ``ValueError``) on any schema deviation so the
    caller can return HTTP 400 BEFORE persisting the file. Performed entirely
    in memory — no disk write on failure.

    Checks:
      1. All ``required_sheets`` exist in the workbook
      2. The ``primary_sheet`` (typically ``"Table_1"``) has a header row at
         ``header_row`` containing at minimum one cell matching EACH key in
         ``required_header_aliases`` (case-insensitive, whitespace-stripped)
      3. At least one non-empty data row below the header
    """
    import io
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        # Pre-validate already ran — getting here means the file passed BadZipFile
        # but can't be re-opened for schema; treat as bad xlsx (client fault).
        raise InvalidExcelSchemaError(
            "Uploaded file could not be parsed as a valid .xlsx workbook."
        ) from e

    # 1. Required sheets
    present = set(wb.sheetnames)
    missing = [s for s in required_sheets if s not in present]
    if missing:
        raise InvalidExcelSchemaError(
            f"Excel file is missing the expected '{missing[0]}' sheet. "
            "Please upload a valid ANZSCO workbook (ABS Feb 2026 format)."
        )

    # 2. Primary sheet header row column check
    ws = wb[primary_sheet]
    # Pull header row cells (read-only workbook iter)
    header_cells: List[Any] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == header_row:
            header_cells = list(row or [])
            break
    headers_lc = {(str(c).strip().lower() if c is not None else "") for c in header_cells}
    headers_lc.discard("")
    for role, aliases in required_header_aliases.items():
        alias_set = {a.strip().lower() for a in aliases}
        if not (headers_lc & alias_set):
            raise InvalidExcelSchemaError(
                f"Excel file is missing the required '{role.capitalize()}' column "
                f"in row {header_row}. Please upload a valid ANZSCO workbook."
            )

    # 3. At least one non-empty data row
    has_data = False
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx <= header_row:
            continue
        if any(c not in (None, "") for c in (row or [])):
            has_data = True
            break
    if not has_data:
        raise InvalidExcelSchemaError(
            "Excel file has no data rows below the header. "
            "Please upload a workbook with at least one occupation."
        )


# ─── Phase 17.0.2 — Orphan-row cleanup ──────────────────────────────────────
async def prune_orphan_failed_rows(db) -> int:
    """Delete any ``import_files`` row whose ``status='failed'`` AND whose
    on-disk artefact no longer exists. Safe to run on every startup.
    Returns the number of rows pruned."""
    pruned = 0
    async for d in db["import_files"].find({"status": "failed"}, {"_id": 0}):
        sp = d.get("storage_path") or ""
        if not sp or not os.path.exists(sp):
            await db["import_files"].delete_one({"id": d.get("id")})
            pruned += 1
    if pruned:
        logger.info("[import_storage] Pruned %d orphan failed rows on startup", pruned)
    return pruned


async def delete_file(db, file_id: str) -> None:
    """Remove a file_files row + its on-disk artefact. Used when a persisted
    upload subsequently fails import-time validation."""
    doc = await db["import_files"].find_one({"id": file_id}, {"_id": 0})
    if not doc:
        return
    sp = doc.get("storage_path")
    if sp:
        try:
            os.unlink(sp)
        except OSError as e:
            logger.warning("[import_storage] Could not unlink %s on rollback: %s", sp, e)
    await db["import_files"].delete_one({"id": file_id})
    # Re-promote the next-most-recent file (if any) to is_latest=True
    next_latest = await db["import_files"].find_one(
        {"source_type": doc.get("source_type")},
        sort=[("uploaded_at", -1)],
    )
    if next_latest:
        await db["import_files"].update_one(
            {"id": next_latest["id"]},
            {"$set": {"is_latest": True}},
        )
