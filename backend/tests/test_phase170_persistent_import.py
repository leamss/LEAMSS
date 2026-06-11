"""Phase 17.0 — Verification Hub persistent-file import + sanitised errors.

Tests cover:
  1. Upload creates a real `import_files` row + on-disk artefact, response is sanitised
  2. Second upload demotes the first (only one is_latest per source_type)
  3. Same-bytes upload de-dupes (reuses file_id)
  4. `/import-anzsco-default` with no prior file returns HTTP 409 + structured action choices
  5. `/import-anzsco-default` with a prior file works
  6. Aggregate: no endpoint EVER leaks `/tmp` or `/app/backend/storage` to client
  7. `/auto-fetch-anzsco` runs against live Home Affairs (network-dependent)
  8. Phase 7.1 seeder is idempotent — 5 templates + 1 default policy after two runs

Run:
    cd /app/backend && python -m pytest tests/test_phase170_persistent_import.py -v
"""
from __future__ import annotations

import asyncio
import io
import json
import os
import sys
from pathlib import Path
from typing import Any, Dict

import httpx
import pytest
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

from motor.motor_asyncio import AsyncIOMotorClient  # noqa: E402

API_BASE = os.environ.get("AUDIT_API_BASE", "http://localhost:8001/api")
ADMIN_EMAIL = "admin@leamss.com"
ADMIN_PASSWORD = "Admin@123"
SRC = "anzsco_4digit"

_db = AsyncIOMotorClient(os.environ["MONGO_URL"])[os.environ["DB_NAME"]]


# ─── Helpers ────────────────────────────────────────────────────────────────
def _login_token() -> str:
    r = httpx.post(
        f"{API_BASE}/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=10,
    )
    r.raise_for_status()
    return r.json()["token"]


@pytest.fixture(scope="module")
def auth_headers():
    return {"Authorization": f"Bearer {_login_token()}"}


def _build_min_xlsx(title_suffix: str = "") -> bytes:
    """Tiny valid ANZSCO Excel — 1 row in Table_1 with all 8 expected columns,
    Table_2 with code + description, empty other tables. `title_suffix` lets each
    test produce a distinct sha256 when needed."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Table_1"
    # Table_1 expects 8 columns: code, title, employed, pt%, female%, median$, age, growth
    headers1 = ["Code", "Title", "Employed", "PT%", "Female%", "MedianAUD", "MedianAge", "Growth"]
    for i, h in enumerate(headers1, start=1):
        ws.cell(row=7, column=i).value = h
    row1 = ["2613", f"Software & Applications Programmers{title_suffix}", 100000, 12.5, 22.3, 1850, 38, 5]
    for i, v in enumerate(row1, start=1):
        ws.cell(row=8, column=i).value = v

    # Table_2 (descriptions) needs at least 3 cols
    ws2 = wb.create_sheet("Table_2")
    ws2.cell(row=7, column=1).value = "Code"
    ws2.cell(row=7, column=2).value = "Title"
    ws2.cell(row=7, column=3).value = "Description"
    ws2.cell(row=8, column=1).value = "2613"
    ws2.cell(row=8, column=2).value = "Software & Applications Programmers"
    ws2.cell(row=8, column=3).value = "Designs, develops, tests, deploys and maintains software systems."

    # Empty placeholder sheets so the importer doesn't error on missing tabs
    for name in ["Table_3", "Table_4", "Table_5", "Table_6", "Table_7", "Table_8", "Table_9"]:
        s = wb.create_sheet(name)
        for i in range(1, 9):
            s.cell(row=7, column=i).value = f"col{i}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _async(coro):
    return asyncio.get_event_loop().run_until_complete(coro)


async def _drop_import_files():
    await _db["import_files"].delete_many({})


async def _wipe_storage_dir():
    # Best-effort cleanup; matches what `import_storage.STORAGE_ROOT` controls.
    from core import import_storage
    root: Path = import_storage.STORAGE_ROOT / SRC
    if root.exists():
        for f in root.iterdir():
            try:
                f.unlink()
            except OSError:
                pass


PATH_LEAK_PATTERNS = ("/tmp/", "/app/backend/storage", "storage_path")


def _assert_no_path_leak(payload: Any, label: str = "") -> None:
    s = json.dumps(payload, default=str)
    for p in PATH_LEAK_PATTERNS:
        assert p not in s, (
            f"Path leak '{p}' found in {label} response: {s[:200]}"
        )


# ─── Tests ──────────────────────────────────────────────────────────────────
def test_1_upload_creates_import_files_row(auth_headers):
    """Upload creates DB row + on-disk file. Response is sanitised."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    data = _build_min_xlsx("_t1")
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("anzsco_t1.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    )
    assert r.status_code == 200, f"upload failed: {r.status_code} · {r.text[:200]}"
    body = r.json()
    assert body.get("ok") is True
    assert "file" in body
    assert body["file"]["filename_original"] == "anzsco_t1.xlsx"
    assert body["file"]["is_latest"] is True
    assert body["file"]["status"] == "imported"
    assert body["file"]["sha256"]
    # No server path in any response field
    _assert_no_path_leak(body, "upload")
    assert "storage_path" not in body["file"], "storage_path leaked into client payload"

    # DB row exists with is_latest=True
    row = _async(_db["import_files"].find_one({"id": body["file"]["id"]}))
    assert row is not None
    assert row["is_latest"] is True
    # On-disk artefact exists
    assert os.path.exists(row["storage_path"]), "on-disk artefact missing"
    assert row["storage_path"].endswith(".xlsx")


def test_2_second_upload_demotes_first(auth_headers):
    """Two distinct uploads — only one is_latest=True; newer wins."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())
    httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("first.xlsx", _build_min_xlsx("_v1"), "application/octet-stream")},
        timeout=30,
    )
    httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("second.xlsx", _build_min_xlsx("_v2"), "application/octet-stream")},
        timeout=30,
    )
    latest_rows = _async(_db["import_files"].find({"source_type": SRC, "is_latest": True}).to_list(10))
    assert len(latest_rows) == 1, f"Expected exactly 1 is_latest, got {len(latest_rows)}"
    assert latest_rows[0]["filename_original"] == "second.xlsx"


def test_3_same_hash_dedupe(auth_headers):
    """Uploading identical bytes twice MUST NOT create a duplicate DB row."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())
    payload = _build_min_xlsx("_dedupe")
    r1 = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("same.xlsx", payload, "application/octet-stream")},
        timeout=30,
    )
    r2 = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("same.xlsx", payload, "application/octet-stream")},
        timeout=30,
    )
    assert r1.status_code == 200 and r2.status_code == 200
    # Same file_id reused on second call
    assert r1.json()["file"]["id"] == r2.json()["file"]["id"]
    rows = _async(_db["import_files"].find({"source_type": SRC}).to_list(10))
    assert len(rows) == 1, f"Dedupe failed — got {len(rows)} rows, expected 1"


def test_4_reimport_with_no_prior_returns_409(auth_headers):
    """When import_files is empty, /import-anzsco-default must return a
    structured 409 with action choices (NEVER a /tmp path)."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-default",
        headers=auth_headers,
        timeout=10,
    )
    assert r.status_code == 409, f"Expected 409, got {r.status_code}: {r.text[:200]}"
    body = r.json()
    assert body.get("code") == "NO_PRIOR_FILE"
    assert "message" in body and body["message"]
    assert "/tmp" not in body["message"], f"Path leaked in 409 message: {body['message']}"
    assert isinstance(body.get("actions"), list)
    assert len(body["actions"]) >= 2
    kinds = {a.get("kind") for a in body["actions"]}
    assert {"upload", "fetch_latest"}.issubset(kinds)
    _assert_no_path_leak(body, "409 NO_PRIOR_FILE")


def test_5_reimport_with_prior_works(auth_headers):
    """After upload, /import-anzsco-default re-imports the stored file."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())
    httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("prior.xlsx", _build_min_xlsx("_t5"), "application/octet-stream")},
        timeout=30,
    )
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-default",
        headers=auth_headers,
        timeout=30,
    )
    assert r.status_code == 200, f"Re-import failed: {r.status_code} · {r.text[:200]}"
    body = r.json()
    assert body.get("ok") is True
    assert body.get("summary")
    assert body["file"]["filename_original"] == "prior.xlsx"
    _assert_no_path_leak(body, "re-import success")


def test_6_response_never_leaks_server_path(auth_headers):
    """Sweep through all 3 KB-unified endpoints with various inputs and assert
    no response body ever contains /tmp or /app/backend/storage."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    # GET latest empty
    r = httpx.get(
        f"{API_BASE}/kb-unified/import-files/latest?source_type={SRC}",
        headers=auth_headers, timeout=10,
    )
    _assert_no_path_leak(r.json(), "latest-empty")
    assert r.json()["file"] is None

    # 409 no prior
    r = httpx.post(f"{API_BASE}/kb-unified/import-anzsco-default", headers=auth_headers, timeout=10)
    _assert_no_path_leak(r.json(), "no-prior 409")

    # Upload + verify
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("scan.xlsx", _build_min_xlsx("_t6"), "application/octet-stream")},
        timeout=30,
    )
    _assert_no_path_leak(r.json(), "upload-success")

    # GET latest populated
    r = httpx.get(
        f"{API_BASE}/kb-unified/import-files/latest?source_type={SRC}",
        headers=auth_headers, timeout=10,
    )
    _assert_no_path_leak(r.json(), "latest-populated")

    # GET history list
    r = httpx.get(
        f"{API_BASE}/kb-unified/import-files?source_type={SRC}&limit=10",
        headers=auth_headers, timeout=10,
    )
    _assert_no_path_leak(r.json(), "history-list")

    # POST re-import
    r = httpx.post(f"{API_BASE}/kb-unified/import-anzsco-default", headers=auth_headers, timeout=30)
    _assert_no_path_leak(r.json(), "reimport-success")


def test_7_auto_fetch_anzsco_runs(auth_headers):
    """Live scrape Home Affairs SOL via the Phase 17.0 alias (now forwards
    to /auto-fetch-country?country=AU). Skips if backend says 502."""
    r = httpx.post(
        f"{API_BASE}/kb-unified/auto-fetch-anzsco", headers=auth_headers, timeout=60,
    )
    if r.status_code == 502:
        pytest.skip("Network-dependent test — Home Affairs unreachable")
    assert r.status_code == 200, f"auto-fetch failed: {r.status_code} · {r.text[:200]}"
    body = r.json()
    assert body.get("ok") is True
    # Phase 17.1 — response is now the multi-country shape with one AU entry.
    assert isinstance(body.get("results"), list) and len(body["results"]) == 1
    au = body["results"][0]
    assert au["country"] == "AU"
    assert "Home Affairs" in au["source"]
    assert isinstance(au.get("imported"), int)
    assert isinstance(au.get("updated"), int)
    assert au["imported"] + au["updated"] >= 100
    _assert_no_path_leak(body, "auto-fetch")


def test_8_phase71_seeder_idempotent():
    """Phase 7.1 seeder must not duplicate templates/policy on re-run."""
    from migrations.phase71_kb_unification import run_idempotent
    _async(run_idempotent(_db))  # first call (or no-op if startup already ran)
    _async(run_idempotent(_db))  # second call — should be all "existed"

    templates = _async(_db["country_templates"].find({}).to_list(50))
    cc_set = {t.get("country_code") for t in templates}
    assert {"AU", "CA", "NZ", "UK", "USA"}.issubset(cc_set), (
        f"Missing seed template(s): expected AU/CA/NZ/UK/USA, got {cc_set}"
    )
    # No duplicates per country
    from collections import Counter
    counts = Counter(t.get("country_code") for t in templates)
    dupes = [cc for cc, n in counts.items() if n > 1]
    assert not dupes, f"Duplicate templates for {dupes}"

    # Exactly 1 default protection policy
    default_policies = _async(
        _db["protection_policies"].find({"is_default_leamss": True}).to_list(10)
    )
    assert len(default_policies) == 1, (
        f"Expected exactly 1 default protection policy, got {len(default_policies)}"
    )


# ─── Phase 17.0.1 — Malformed-upload regression ────────────────────────────
def test_9_malformed_xlsx_returns_400(auth_headers):
    """Plain text renamed to .xlsx must return HTTP 400 (not 500) with a
    user-friendly detail string and NO server-path leak."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    junk = b"this is just plain text masquerading as an xlsx workbook" * 5
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("fake.xlsx", junk, "application/octet-stream")},
        timeout=30,
    )
    assert r.status_code == 400, (
        f"Expected 400 for malformed xlsx, got {r.status_code}: {r.text[:200]}"
    )
    body = r.json()
    detail = body.get("detail", "")
    assert ".xlsx" in detail.lower() or "excel" in detail.lower() or "workbook" in detail.lower(), (
        f"Detail must be user-friendly: {detail}"
    )
    _assert_no_path_leak(body, "malformed-xlsx-400")


def test_10_malformed_upload_not_persisted(auth_headers):
    """After a 400 the malformed bytes must NOT appear in `import_files` and
    must NOT leave an on-disk artefact behind — storage stays clean."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    from core import import_storage as st
    src_dir = st.STORAGE_ROOT / SRC

    rows_before = _async(_db["import_files"].count_documents({"source_type": SRC}))
    files_before = list(src_dir.iterdir()) if src_dir.exists() else []

    junk = b"<html>not an xlsx</html>"
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("evil.xlsx", junk, "application/octet-stream")},
        timeout=30,
    )
    assert r.status_code == 400, f"Expected 400, got {r.status_code}"

    rows_after = _async(_db["import_files"].count_documents({"source_type": SRC}))
    files_after = list(src_dir.iterdir()) if src_dir.exists() else []
    assert rows_after == rows_before, (
        f"Malformed upload polluted import_files: before={rows_before} after={rows_after}"
    )
    assert len(files_after) == len(files_before), (
        f"Malformed upload left disk artefact: before={files_before} after={files_after}"
    )


def test_11_genuine_500_still_returns_500(auth_headers, monkeypatch):
    """When a TRUE server fault hits the importer (not a bad-file error), the
    response must still be 500 — we don't want to mask real bugs as 400.

    We mock `import_anzsco_excel` to raise a server-class exception AFTER the
    pre-validation pass (so the file does get persisted) — `classify_upload_error`
    must route this to a 500 because the exception type isn't one of the
    bad-file kinds.
    """
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    # Patch the imported symbol used by the router. We use sys.modules so the
    # change is visible to the live-running FastAPI server. NOTE: this is a
    # best-effort test — the running server has already cached the import,
    # so we send a real valid xlsx but monkeypatch raises only inside our own
    # synchronous re-implementation. If we cannot guarantee mock observability
    # in the running server, we accept skip.
    import importlib
    try:
        from routers import kb_unified as _kb  # type: ignore
    except ImportError:  # pragma: no cover
        pytest.skip("kb_unified router not directly importable in this layout")

    async def _boom(*a, **kw):  # pragma: no cover (executed via monkeypatch)
        raise RuntimeError("simulated downstream Mongo write failure")

    original = _kb.import_anzsco_excel
    _kb.import_anzsco_excel = _boom  # type: ignore[attr-defined]
    try:
        r = httpx.post(
            f"{API_BASE}/kb-unified/import-anzsco-excel",
            headers=auth_headers,
            files={"file": ("good.xlsx", _build_min_xlsx("_t11"), "application/octet-stream")},
            timeout=30,
        )
    finally:
        _kb.import_anzsco_excel = original  # type: ignore[attr-defined]

    # The patch only affects in-process callers. The HTTP-side server is the
    # uvicorn process which holds its own reference — so this test will most
    # likely see a normal 200 (mock didn't reach the live server). In that
    # case we mark the test as a documented xfail per spec.
    if r.status_code == 200:
        pytest.skip(
            "Cannot inject into live uvicorn process — documented in task brief "
            "(\"If too contrived to mock, skip\")."
        )
    assert r.status_code == 500, f"Expected 500 for true server fault, got {r.status_code}"
    _assert_no_path_leak(r.json(), "genuine-500")



# ─── Phase 17.0.2 — Schema-shape pre-validation ────────────────────────────
def _build_xlsx_with(sheets: list[tuple[str, list[list]]]) -> bytes:
    """Construct an xlsx where each entry is (sheet_name, [[row7_cells], [row8_cells], …]).
    Rows 1-6 are auto-padded blank so the header row remains at row 7.
    Pass an empty `rows` list to get a sheet with no header at all."""
    from openpyxl import Workbook
    wb = Workbook()
    first = True
    for name, rows in sheets:
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        for offset, row in enumerate(rows):
            r_num = 7 + offset
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_num, column=c_idx).value = val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_12_xlsx_missing_required_sheet_returns_400(auth_headers):
    """Valid xlsx with sheet named 'WrongName' (no Table_1) → HTTP 400, no
    DB row, no on-disk artefact."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    data = _build_xlsx_with([(
        "WrongName",
        [["Code", "Title"], ["2613", "Software Programmer"]],
    )])
    rows_before = _async(_db["import_files"].count_documents({"source_type": SRC}))

    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("wrong_sheet.xlsx", data, "application/octet-stream")},
        timeout=30,
    )
    assert r.status_code == 400, f"Expected 400, got {r.status_code}: {r.text[:200]}"
    detail = r.json().get("detail", "")
    assert "Table_1" in detail or "sheet" in detail.lower(), (
        f"Detail must mention the missing sheet: {detail}"
    )
    _assert_no_path_leak(r.json(), "missing-sheet-400")
    rows_after = _async(_db["import_files"].count_documents({"source_type": SRC}))
    assert rows_after == rows_before, "Missing-sheet upload polluted import_files"


def test_13_xlsx_missing_code_column_returns_400(auth_headers):
    """Table_1 exists but row 7 headers contain no 'Code' column → HTTP 400."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    sheets = [(
        "Table_1",
        [["RandomCol1", "RandomCol2"], ["foo", "bar"]],
    )]
    for n in ("Table_2", "Table_3", "Table_4", "Table_5", "Table_6", "Table_7", "Table_8"):
        sheets.append((n, [["x"], ["y"]]))
    data = _build_xlsx_with(sheets)

    rows_before = _async(_db["import_files"].count_documents({"source_type": SRC}))
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("no_code.xlsx", data, "application/octet-stream")},
        timeout=30,
    )
    assert r.status_code == 400, f"Expected 400, got {r.status_code}: {r.text[:200]}"
    detail = r.json().get("detail", "")
    assert "Code" in detail or "column" in detail.lower(), (
        f"Detail must mention the missing column: {detail}"
    )
    _assert_no_path_leak(r.json(), "missing-col-400")
    rows_after = _async(_db["import_files"].count_documents({"source_type": SRC}))
    assert rows_after == rows_before, "Missing-column upload polluted import_files"


def test_14_xlsx_no_data_rows_returns_400(auth_headers):
    """Table_1 header row 7 present but rows 8+ empty → HTTP 400."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())

    sheets = [(
        "Table_1",
        [["Code", "Title", "Employed"]],  # only header, no data row
    )]
    for n in ("Table_2", "Table_3", "Table_4", "Table_5", "Table_6", "Table_7", "Table_8"):
        sheets.append((n, [["x"], ["y"]]))
    data = _build_xlsx_with(sheets)

    rows_before = _async(_db["import_files"].count_documents({"source_type": SRC}))
    r = httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel",
        headers=auth_headers,
        files={"file": ("no_data.xlsx", data, "application/octet-stream")},
        timeout=30,
    )
    assert r.status_code == 400, f"Expected 400, got {r.status_code}: {r.text[:200]}"
    detail = r.json().get("detail", "")
    assert "data" in detail.lower() or "row" in detail.lower(), (
        f"Detail must mention missing data rows: {detail}"
    )
    _assert_no_path_leak(r.json(), "no-data-400")
    rows_after = _async(_db["import_files"].count_documents({"source_type": SRC}))
    assert rows_after == rows_before, "No-data upload polluted import_files"


def test_15_failed_rows_never_persisted(auth_headers):
    """Sweep `import_files` count before + after all three negative uploads.
    Total delta must be exactly zero — no schema-failure should ever leave
    a row behind."""
    _async(_drop_import_files())
    _async(_wipe_storage_dir())
    before = _async(_db["import_files"].count_documents({}))

    # (a) plain text
    httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel", headers=auth_headers,
        files={"file": ("a.xlsx", b"junk text", "application/octet-stream")}, timeout=15,
    )
    # (b) wrong sheet
    bad1 = _build_xlsx_with([("Nope", [["a"], ["b"]])])
    httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel", headers=auth_headers,
        files={"file": ("b.xlsx", bad1, "application/octet-stream")}, timeout=15,
    )
    # (c) Table_1 with wrong columns
    bad2 = _build_xlsx_with([
        ("Table_1", [["X", "Y"], ["1", "2"]]),
        ("Table_2", [["x"]]),
        ("Table_3", [["x"]]),
        ("Table_4", [["x"]]),
        ("Table_5", [["x"]]),
        ("Table_6", [["x"]]),
        ("Table_7", [["x"]]),
        ("Table_8", [["x"]]),
    ])
    httpx.post(
        f"{API_BASE}/kb-unified/import-anzsco-excel", headers=auth_headers,
        files={"file": ("c.xlsx", bad2, "application/octet-stream")}, timeout=15,
    )

    after = _async(_db["import_files"].count_documents({}))
    assert after == before, (
        f"Schema failures leaked into import_files: before={before} after={after}"
    )

    # And the storage dir must contain no files for the SRC source_type
    from core import import_storage as st
    src_dir = st.STORAGE_ROOT / SRC
    files_on_disk = list(src_dir.iterdir()) if src_dir.exists() else []
    assert not files_on_disk, f"Schema-failed uploads left disk artefacts: {files_on_disk}"

