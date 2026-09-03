"""
Atlas Industry parser.

Parses Jobs & Skills Atlas Industry export workbook.

This parser is for Atlas exports and DOES NOT replace
industry_data.py which supports the legacy JSA workbook.
"""

from typing import Dict, Any, Iterator
from openpyxl import load_workbook
from datetime import datetime, timezone

def parse_workbook(path: str) -> Iterator[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)

    metadata = _read_contents_sheet(wb)

    industry = {
        "anzsic_code": None,
        "industry_name": None,
        "slug": None,
        "quarterly_time_series": [],
        "employment_projections": [],
        "top_regions": [],
        "top_occupations": [],
        "median_weekly_earnings": [],
        "business": [],
        "industry_subdivisions": [],
        "reos_time_series": [],
        "source": "Jobs and Skills Atlas",
        "source_url": metadata.get("Source URL", ""),
        "latest_update": metadata.get("Latest update"),
        "download_date": metadata.get("Download date"),
        "last_imported_at": datetime.now(timezone.utc).isoformat(),
    }
    ws = wb["Quarterly Time series"]

    first = True

    for row in ws.iter_rows(min_row=2, values_only=True):

        if first:
            industry["anzsic_code"] = row[0]
            industry["industry_name"] = row[1]
            industry["slug"] = (
    row[1]
    .lower()
    .replace("&", "and")
    .replace(",", "")
    .replace("/", "-")
    .replace(" ", "-")
)
            first = False

        industry["quarterly_time_series"].append({
            "quarter_ending": row[4],
            "headcount": row[5],
            "vacancies": row[6]
        })
    ws = wb["Employment Projections"]

    for row in ws.iter_rows(min_row=2, values_only=True):

        industry["employment_projections"].append({
            "date": row[4],
            "projected_employment": row[5]
        })
    ws = wb["Top 10 Regions"]

    for row in ws.iter_rows(min_row=2, values_only=True):

        industry["top_regions"].append({
            "rank": row[5],
            "region_code": row[6],
            "region_name": row[7],
            "headcount": row[8]
        })
    ws = wb["Top 10 Occupations"]

    for row in ws.iter_rows(min_row=2, values_only=True):

        industry["top_occupations"].append({
        "rank": row[5],
        "anzsco_4digit": row[6],
        "occupation_name": row[7],
        "headcount": row[8]
      })
    ws = wb["Median weekly earnings"]

    for row in ws.iter_rows(min_row=2, values_only=True):

        industry["median_weekly_earnings"].append({
            "survey_month": row[4],
            "median_weekly_earnings": row[5],
            "female": row[6],
            "male": row[7]
        })
    ws = wb["Business"]

    for row in ws.iter_rows(min_row=2, values_only=True):

        industry["business"].append({
    "month_beginning": row[4],
    "active_businesses": row[5],
    "entries": row[6],
    "exits": row[7]
})
    ws = wb["Industry subdivisions"]

    for row in ws.iter_rows(min_row=2, values_only=True):

      industry["industry_subdivisions"].append({
        "anzsic_code": row[0],
        "name": row[1],
        "level": row[3],
        "headcount": row[4]
    })
    ws = wb["REOS time series"]

    for row in ws.iter_rows(min_row=2, values_only=True):

       if row[0] is None:
        continue

       industry["reos_time_series"].append({
        "series_date": row[4],
        "recruitment_rate": row[5],
        "recruitment_difficulty_rate": row[6],
        "expect_staff_increase": row[7]
    })
    wb.close()

    yield industry
def _read_contents_sheet(wb):
    """
    Read metadata from the Contents sheet.
    """
    if "Contents" not in wb.sheetnames:
        return {}

    ws = wb["Contents"]

    metadata = {}

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue

        key = str(row[0]).strip() if row[0] else ""
        value = row[1]

        if key:
            metadata[key] = value

    return metadata


def parse_summary(path: str) -> Dict[str, Any]:

    records = list(parse_workbook(path))

    return {
        "source": "Jobs and Skills Atlas Industry Export",
        "row_count": len(records),
        "sample": records[:1],
    }