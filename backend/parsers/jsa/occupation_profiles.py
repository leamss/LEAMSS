"""Phase 19.4 — JSA Occupation Profiles (Feb 2026) parser.

Parses the 10-sheet workbook produced by Jobs and Skills Australia and yields
one `OccupationProfile` dict per **4-digit ANZSCO code**. Mapping to 6-digit
codes (which is what `occupation_master` stores) is done by the importer with
explicit `_parent_inherited: True` flagging on each inherited record.

The data shape is documented in Sub-task B of the Phase 19.4 brief.
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, Iterator

from openpyxl import load_workbook

SOURCE_NAME = "JSA Occupation Profiles Feb 2026"
SOURCE_URL = "https://www.jobsandskills.gov.au/data/occupation-and-industry-profiles"
DATA_QUALITY = "official_govt_data"

# Table 1 header row index (0-based) — verified manually
HEADER_ROWS = {
    "Table_1": 6, "Table_4": 6, "Table_5": 6, "Table_6": 6,
    "Table_7": 6, "Table_8": 6,
}


def _safe_int(v: Any) -> int | None:
    if v in (None, "", "N/A", "n/a"):
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _safe_float(v: Any) -> float | None:
    if v in (None, "", "N/A", "n/a"):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_workbook(path: str) -> Iterator[Dict[str, Any]]:
    """Yield one dict per 4-digit ANZSCO code with all 9 tables joined.

    Each record carries: anzsco_4digit, occupation_title, abs_data {...},
    source, source_url, last_imported_at, data_quality.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    # build per-code accumulator
    by_code: Dict[int, Dict[str, Any]] = {}

    # ── Table 1 — overview (employed, female %, hourly, median weekly, age, growth) ──
    ws = wb["Table_1"]
    for row in ws.iter_rows(values_only=True, min_row=8):
        code = _safe_int(row[0])
        if not code:
            continue
        rec = by_code.setdefault(code, {"anzsco_4digit": str(code).zfill(4), "abs_data": {}})
        rec["occupation_title"] = (row[1] or "").strip() if row[1] else ""
        abs_d = rec["abs_data"]
        abs_d["employed_count"] = _safe_int(row[2])
        abs_d["part_time_share_pct"] = _safe_float(row[3])
        abs_d["female_share_pct"] = _safe_float(row[4])
        abs_d["median_weekly_earnings_aud"] = _safe_float(row[5])
        abs_d["median_age"] = _safe_int(row[6])
        abs_d["annual_employment_growth"] = _safe_int(row[7])

    # ── Table 4 — earnings details ──
    ws = wb["Table_4"]
    for row in ws.iter_rows(values_only=True, min_row=8):
        code = _safe_int(row[0])
        if not code or code not in by_code:
            continue
        abs_d = by_code[code]["abs_data"]
        abs_d["ft_share_pct"] = _safe_float(row[2])
        abs_d["avg_ft_hours_per_week"] = _safe_float(row[3])
        abs_d["median_ft_weekly_earnings_aud"] = _safe_float(row[4])
        abs_d["median_ft_hourly_earnings_aud"] = _safe_float(row[5])
        wkly = abs_d.get("median_ft_weekly_earnings_aud")
        if wkly:
            abs_d["median_ft_annual_aud"] = round(wkly * 52)

    # ── Table 5 — top industries (multi-row) ──
    ws = wb["Table_5"]
    for row in ws.iter_rows(values_only=True, min_row=8):
        code = _safe_int(row[0])
        ind = row[2] if len(row) > 2 else None
        if not code or code not in by_code or not ind:
            continue
        abs_d = by_code[code]["abs_data"]
        inds = abs_d.setdefault("top_industries", [])
        if len(inds) < 5:  # top 5 only
            inds.append({"name": str(ind).strip()})

    # ── Table 6 — state distribution ──
    ws = wb["Table_6"]
    states = ["NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT"]
    for row in ws.iter_rows(values_only=True, min_row=8):
        code = _safe_int(row[0])
        if not code or code not in by_code:
            continue
        abs_d = by_code[code]["abs_data"]
        sd = {}
        for i, s in enumerate(states):
            v = _safe_float(row[2 + i]) if (2 + i) < len(row) else None
            if v is not None:
                sd[s] = v
        if sd:
            abs_d["state_distribution"] = sd

    # ── Table 7 — age profile ──
    ws = wb["Table_7"]
    age_bands = ["15-19", "20-24", "25-34", "35-44", "45-54", "55-59", "60-64", "65+"]
    for row in ws.iter_rows(values_only=True, min_row=8):
        code = _safe_int(row[0])
        if not code or code not in by_code:
            continue
        abs_d = by_code[code]["abs_data"]
        ap = {}
        for i, b in enumerate(age_bands):
            v = _safe_float(row[2 + i]) if (2 + i) < len(row) else None
            if v is not None:
                ap[b] = v
        if ap:
            abs_d["age_profile"] = ap

    # ── Table 8 — education attainment ──
    ws = wb["Table_8"]
    edu_keys = [
        "postgrad_pct", "bachelor_pct", "diploma_pct",
        "certIII_IV_pct", "year12_pct", "year11_pct", "year10_below_pct",
    ]
    for row in ws.iter_rows(values_only=True, min_row=8):
        code = _safe_int(row[0])
        if not code or code not in by_code:
            continue
        abs_d = by_code[code]["abs_data"]
        ed = {}
        for i, k in enumerate(edu_keys):
            v = _safe_float(row[2 + i]) if (2 + i) < len(row) else None
            if v is not None:
                ed[k] = v
        if ed:
            abs_d["education_attainment"] = ed

    wb.close()

    now = datetime.now(timezone.utc).isoformat()
    for code, rec in by_code.items():
        rec["abs_data"]["source"] = SOURCE_NAME
        rec["abs_data"]["source_url"] = SOURCE_URL
        rec["abs_data"]["last_imported_at"] = now
        rec["abs_data"]["data_quality"] = DATA_QUALITY
        yield rec


def parse_summary(path: str) -> Dict[str, Any]:
    """Return a 5-row preview + total count without yielding everything."""
    recs = list(parse_workbook(path))
    return {
        "source": SOURCE_NAME,
        "source_url": SOURCE_URL,
        "row_count": len(recs),
        "sample": recs[:5],
    }
