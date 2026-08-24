from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, List

from openpyxl import load_workbook
from datetime import datetime

SOURCE_NAME = "JSA Atlas Occupation Workbook"
SOURCE_URL = "https://www.jobsandskills.gov.au/data/occupation-and-industry-profiles"
DATA_QUALITY = "official_govt_data"

def parse_workbook(path: str):

    wb = load_workbook(path, read_only=True, data_only=True)

    result = {}

    result["metadata"] = _parse_contents(wb["Contents"])

    result["monthly_time_series"] = parse_monthly_time_series(
        wb["Monthly Time series"]
    )

    result["quarterly_time_series"] = parse_quarterly_time_series(
        wb["Quarterly Time series"]
    )

    result["employer_recruitment_insights"] = parse_employer_recruitment_insights(
        wb["Employer recruitment insights"]
    )

    result["top10s"] = parse_top10s(
        wb["Top 10s"]
    )

    result["demographic_data"] = parse_demographic_data(
        wb["Demographic data"]
    )

    result["main_fields_of_education"] = parse_main_fields_of_education(
        wb["Main fields of education"]
    )

    result["shortage_ratings"] = parse_shortage_ratings(
        wb["Shortage ratings"]
    )

    result["projected_employment"] = parse_projected_employment(
        wb["Projected employment"]
    )

    result["apprentice_trainee_time_series"] = parse_apprentice_trainee_time_series(
        wb["A&T time series"]
    )

    result["apprentice_trainee_demographics"] = parse_apprentice_trainee_demographics(
        wb["A&T demographics"]
    )

    result["occupational_mobility"] = parse_occupational_mobility(
        wb["Occupational Mobility"]
    )

    wb.close()

    yield result


def _clean(v):

    if v is None:
        return None

    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")

    if isinstance(v, str):
        v = v.strip()
        if not v:
            return None

    return v

def parse_simple_table(ws) -> List[Dict[str, Any]]:
    """
    Generic parser for JSA Atlas sheets.
    Assumes:
      Row 1 = headers
      Row 2 onwards = data
    """

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [_clean(h) for h in rows[0]]

    records = []

    for row in rows[1:]:

        if not any(row):
            continue

        record = {}

        for header, value in zip(headers, row):

            if header is None:
                continue

            record[str(header)] = _clean(value)

        records.append(record)

    return records
def _parse_contents(ws):

    meta = {}

    for row in ws.iter_rows(values_only=True):

        if len(row) < 3:
            continue

        key = _clean(row[1])
        value = _clean(row[2])

        if key:
            meta[key] = value

    return meta
def parse_monthly_time_series(ws):
    return parse_simple_table(ws)


def parse_quarterly_time_series(ws):
    return parse_simple_table(ws)


def parse_employer_recruitment_insights(ws):
    return parse_simple_table(ws)


def parse_top10s(ws):
    return parse_simple_table(ws)


def parse_demographic_data(ws):
    return parse_simple_table(ws)


def parse_main_fields_of_education(ws):
    return parse_simple_table(ws)


def parse_shortage_ratings(ws):
    return parse_simple_table(ws)


def parse_projected_employment(ws):
    return parse_simple_table(ws)


def parse_apprentice_trainee_time_series(ws):
    return parse_simple_table(ws)


def parse_apprentice_trainee_demographics(ws):
    return parse_simple_table(ws)


def parse_occupational_mobility(ws):
    return parse_simple_table(ws)

def parse_summary(path: str):
    """
    Preview shown in Universal Import Hub before import.
    """

    records = list(parse_workbook(path))

    return {
     "source": SOURCE_NAME,
     "source_url": SOURCE_URL,
     "row_count": len(records),
     "sample": records[:1],
}