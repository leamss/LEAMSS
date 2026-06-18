"""Phase 19.4 — JSA Employment Projections (May 2025–May 2035) parser.

Yields per-4-digit-ANZSCO record with `jsa_data` block. We use Table_6
("Occupation Unit Group") which is the 4-digit ANZSCO level.

Growth-category mapping (per Sir's Sub-task C spec):
  >20%  → "Very Strong"
  10-20% → "Strong"
  0-10%  → "Moderate"
  -5-0%  → "Stable"
  <-5%   → "Declining"
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, Iterator

from openpyxl import load_workbook

SOURCE_NAME = "JSA Employment Projections May 2025-2035"
SOURCE_URL = "https://www.jobsandskills.gov.au/data/employment-projections"
DATA_QUALITY = "official_govt_data"


def _safe_int(v: Any) -> int | None:
    if v in (None, "", "N/A", "n/a", "-"):
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _safe_float(v: Any) -> float | None:
    if v in (None, "", "N/A", "n/a", "-"):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _classify_growth(pct_10y: float | None) -> str:
    """Map 10-year growth % to categorical pill. Input is FRACTION (0.18 = 18%)."""
    if pct_10y is None:
        return "Unknown"
    p = pct_10y * 100  # to percent
    if p > 20:
        return "Very Strong"
    if p >= 10:
        return "Strong"
    if p >= 0:
        return "Moderate"
    if p >= -5:
        return "Stable"
    return "Declining"


def parse_workbook(path: str) -> Iterator[Dict[str, Any]]:
    """Yield one dict per 4-digit ANZSCO with jsa_data block."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table_6 Occupation Unit Group"]
    now = datetime.now(timezone.utc).isoformat()

    # Header at row 8 (0-indexed 7); data starts row 9 (0-indexed 8)
    # Columns (per inspection):
    # 0: Occupation Level (1=Major, 2=Sub-major, 3=Minor, 4=Unit Group)
    # 1: NFD Indicator (Y/N)
    # 2: ANZSCO Code
    # 3: Occupation
    # 4: Skill level
    # 5: Baseline May 2025 ('000)
    # 6: Projected May 2030 ('000)
    # 7: Projected May 2035 ('000)
    # 8: 5-Year Change Level
    # 9: 5-Year Change %
    # 10: 10-Year Change Level
    # 11: 10-Year Change %

    for row in ws.iter_rows(values_only=True, min_row=10):
        if not row or row[0] is None:
            continue
        try:
            level = int(row[0])
        except (ValueError, TypeError):
            continue
        # Only 4-digit (Unit Group) rows
        if level != 4:
            continue
        # Skip NFD ("Not Further Defined") rows — these are placeholders
        if row[1] == "Y":
            continue
        code = _safe_int(row[2])
        if not code or code < 1000 or code > 9999:
            continue

        base_2025 = _safe_float(row[5])
        proj_2030 = _safe_float(row[6])
        proj_2035 = _safe_float(row[7])
        growth_5y_pct = _safe_float(row[9])  # fraction (0.07 = 7%)
        growth_10y_pct = _safe_float(row[11])

        # Numbers are in '000 → multiply by 1000 for absolute count
        emp_2025 = int(base_2025 * 1000) if base_2025 is not None else None
        emp_2030 = int(proj_2030 * 1000) if proj_2030 is not None else None
        emp_2035 = int(proj_2035 * 1000) if proj_2035 is not None else None

        jsa_data = {
            "employment_2025": emp_2025,
            "projected_employment_2030": emp_2030,
            "projected_employment_2035": emp_2035,
            "growth_pct_2025_to_2030": round(growth_5y_pct * 100, 1) if growth_5y_pct is not None else None,
            "growth_pct_2025_to_2035": round(growth_10y_pct * 100, 1) if growth_10y_pct is not None else None,
            "future_growth": _classify_growth(growth_10y_pct),
            "source": SOURCE_NAME,
            "source_url": SOURCE_URL,
            "last_imported_at": now,
            "data_quality": DATA_QUALITY,
        }

        yield {
            "anzsco_4digit": str(code).zfill(4),
            "occupation_title": (row[3] or "").strip() if row[3] else "",
            "jsa_data": jsa_data,
        }

    wb.close()


def parse_summary(path: str) -> Dict[str, Any]:
    recs = list(parse_workbook(path))
    return {
        "source": SOURCE_NAME,
        "source_url": SOURCE_URL,
        "row_count": len(recs),
        "sample": recs[:5],
    }
