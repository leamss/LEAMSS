"""Phase 19.4 — JSA SA4 Regional Labour Market Ratings parser.

HONEST DISCLOSURE: This file is **region-level**, NOT per-occupation. It contains
~94 SA4 (ABS Statistical Area Level 4) regions with overall labour-market
strength ratings + supporting indicators. There is NO ANZSCO×SA4 breakdown in
the data Sir uploaded; surfacing per-occupation regional demand would be fake.

We therefore model this as `regional_labour_market` — a 94-row collection of
regional strength indicators that the atlas pages can show as "Strong job
markets in this country: Sydney - Eastern Suburbs · ..." rather than claiming
occupation-specific demand.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from typing import Any, Dict, Iterator

from openpyxl import load_workbook

SOURCE_NAME = "JSA Regional Labour Market Indicator (RLMI) — March 2026"
SOURCE_URL = "https://www.jobsandskills.gov.au/data/labour-market-update"
DATA_QUALITY = "official_govt_data"


def _safe_float(v: Any) -> float | None:
    if v in (None, "", "N/A", "n/a", "-"):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# Derive state from SA4 code: NSW 1xx, VIC 2xx, QLD 3xx, SA 4xx, WA 5xx, TAS 6xx, NT 7xx, ACT 8xx
_STATE_MAP = {1: "NSW", 2: "VIC", 3: "QLD", 4: "SA", 5: "WA", 6: "TAS", 7: "NT", 8: "ACT"}


def _state_from_sa4(sa4_code: int) -> str:
    first = sa4_code // 100
    return _STATE_MAP.get(first, "UNKNOWN")


def parse_workbook(path: str) -> Iterator[Dict[str, Any]]:
    """Yield one dict per SA4 region from the latest snapshot sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    # Use the first 'March YYYY' sheet (latest snapshot)
    snapshot_sheet = next((s for s in wb.sheetnames if "March 2026" in s or "March 2025" in s), None)
    if snapshot_sheet is None:
        snapshot_sheet = wb.sheetnames[1]  # fallback to 2nd sheet
    ws = wb[snapshot_sheet]
    now = datetime.now(timezone.utc).isoformat()

    for row in ws.iter_rows(values_only=True, min_row=10):
        if not row or row[0] is None:
            continue
        try:
            sa4_code = int(row[0])
        except (ValueError, TypeError):
            continue
        sa4_name = (row[1] or "").strip() if row[1] else ""
        rating = (row[2] or "").strip() if row[2] else ""
        if not sa4_name or not rating:
            continue

        yield {
            "id": str(uuid.uuid4()),
            "sa4_code": sa4_code,
            "sa4_name": sa4_name,
            "state": _state_from_sa4(sa4_code),
            "rating": rating,
            "indicators": {
                "employment_rate_15_64_pct": _safe_float(row[3]),
                "unemployment_rate_pct": _safe_float(row[4]),
                "jobseeker_support_pct": _safe_float(row[5]),
                "underemployment_rate_pct": _safe_float(row[7]) if len(row) > 7 else None,
                "vacancy_rate_pct": _safe_float(row[8]) if len(row) > 8 else None,
                "vacancy_fill_rate_pct": _safe_float(row[10]) if len(row) > 10 else None,
            },
            "snapshot_period": snapshot_sheet,
            "source": SOURCE_NAME,
            "source_url": SOURCE_URL,
            "last_imported_at": now,
            "data_quality": DATA_QUALITY,
        }

    wb.close()


def parse_summary(path: str) -> Dict[str, Any]:
    recs = list(parse_workbook(path))
    return {
        "source": SOURCE_NAME,
        "source_url": SOURCE_URL,
        "row_count": len(recs),
        "sample": recs[:5],
        "honest_note": (
            "This is SA4 region-level labour-market strength (94 regions), "
            "NOT per-occupation demand. No ANZSCO×SA4 cross-tab exists in this file."
        ),
    }
