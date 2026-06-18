"""Phase 19.4c — JSA Industry Data (Feb 2026) parser.

Parses the 5-sheet workbook (Contents + Table_1..4) into one `IndustryProfile`
record per ANZSIC industry (19 records). Mapped to a new `industry_master`
collection — the 2-way occupation↔industry SEO graph.

Tables consumed:
  Table_1 — Overview (employment, female %, part-time %, median weekly earnings, etc.)
  Table_2 — 20-year quarterly employment history
  Table_4 — Top employing occupations per industry (industry → ANZSCO 4-digit, ranked)
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, Iterator

from openpyxl import load_workbook

SOURCE_NAME = "JSA Industry Data Feb 2026"
SOURCE_URL = "https://www.jobsandskills.gov.au/data/occupation-and-industry-profiles"
DATA_QUALITY = "official_govt_data"

# ANZSIC 1-letter division codes for the 19 industries in JSA's standard ordering.
# Order matches Table_1 of the JSA Industry Data Feb 2026 workbook.
_ANZSIC_DIVISIONS = {
    "Agriculture, Forestry and Fishing": "A",
    "Mining": "B",
    "Manufacturing": "C",
    "Electricity, Gas, Water and Waste Services": "D",
    "Construction": "E",
    "Wholesale Trade": "F",
    "Retail Trade": "G",
    "Accommodation and Food Services": "H",
    "Transport, Postal and Warehousing": "I",
    "Information Media and Telecommunications": "J",
    "Financial and Insurance Services": "K",
    "Rental, Hiring and Real Estate Services": "L",
    "Professional, Scientific and Technical Services": "M",
    "Administrative and Support Services": "N",
    "Public Administration and Safety": "O",
    "Education and Training": "P",
    "Health Care and Social Assistance": "Q",
    "Arts and Recreation Services": "R",
    "Other Services": "S",
}


def _slugify(name: str) -> str:
    """URL-safe slug for atlas industry hub pages."""
    s = name.lower()
    for ch in ",.;:()/&'":
        s = s.replace(ch, "")
    s = "-".join(s.split())
    return s


def _safe_int(v: Any) -> int | None:
    if v in (None, "", "N/A", "n/a", "-"):
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _safe_float(v: Any) -> float | None:
    if v in (None, "", "N/A", "n/a", "-"):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_workbook(path: str) -> Iterator[Dict[str, Any]]:
    """Yield one dict per ANZSIC industry — Table 1 + Table 2 + Table 4 joined."""
    wb = load_workbook(path, read_only=True, data_only=True)
    by_name: Dict[str, Dict[str, Any]] = {}

    # ── Table_1 — Overview ──
    ws = wb["Table_1"]
    # Header at row 6 (0-indexed): Industry, Employed, Female Share (%), Part-time Share (%),
    #                              Median Weekly Earnings, Workforce Share (%), Median Age
    for row in ws.iter_rows(values_only=True, min_row=8):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name not in _ANZSIC_DIVISIONS:
            # Skip "Total" / footer rows
            continue
        wkly = _safe_int(row[4])
        rec = {
            "industry_name": name,
            "anzsic_code": _ANZSIC_DIVISIONS[name],
            "slug": _slugify(name),
            "employed_count": _safe_int(row[1]),
            "female_share_pct": _safe_float(row[2]),
            "part_time_share_pct": _safe_float(row[3]),
            "median_weekly_earnings_aud": wkly,
            "median_ft_annual_aud": (wkly * 52) if wkly else None,
            "workforce_share_pct": _safe_float(row[5]),
            "median_age": _safe_int(row[6]),
            "top_employing_occupations": [],
            "employment_history_20y": [],
        }
        by_name[name] = rec

    # ── Table_2 — 20y employment history ──
    ws = wb["Table_2"]
    # Row 6 has periods (Feb-06, May-06, ...). Data starts row 8.
    period_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=8)):
        if i == 6:
            period_row = list(row)
            break
    if period_row:
        periods = [p for p in period_row[1:] if p]  # skip industry-name column
        for row in ws.iter_rows(values_only=True, min_row=8):
            if not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            if name not in by_name:
                continue
            values = list(row[1:1 + len(periods)])
            by_name[name]["employment_history_20y"] = [
                {"period": str(p), "employed": _safe_int(v)}
                for p, v in zip(periods, values)
                if _safe_int(v) is not None
            ]

    # ── Table_4 — Top employing occupations per industry (industry → ANZSCO 4-digit) ──
    ws = wb["Table_4"]
    # Header row 6: Industry, ANZSCO Code, Occupation (ranked)
    for row in ws.iter_rows(values_only=True, min_row=8):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name not in by_name:
            continue
        code = _safe_int(row[1])
        title = (row[2] or "").strip() if row[2] else ""
        if not code or not title:
            continue
        occ_list = by_name[name]["top_employing_occupations"]
        if len(occ_list) < 10:  # top 10 per industry
            occ_list.append({"anzsco_4digit": str(code).zfill(4), "title": title})

    wb.close()

    now = datetime.now(timezone.utc).isoformat()
    for name, rec in by_name.items():
        rec["source"] = SOURCE_NAME
        rec["source_url"] = SOURCE_URL
        rec["data_quality"] = DATA_QUALITY
        rec["last_imported_at"] = now
        yield rec


def parse_summary(path: str) -> Dict[str, Any]:
    recs = list(parse_workbook(path))
    return {
        "source": SOURCE_NAME,
        "source_url": SOURCE_URL,
        "row_count": len(recs),
        "sample": [{
            "industry_name": r["industry_name"],
            "anzsic_code": r["anzsic_code"],
            "employed_count": r["employed_count"],
            "median_weekly_earnings_aud": r["median_weekly_earnings_aud"],
            "top_3_occupations": [o["title"] for o in r["top_employing_occupations"][:3]],
        } for r in recs[:5]],
    }
