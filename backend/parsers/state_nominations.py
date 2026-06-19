"""Phase 19.10 — State Nomination List Parser.

Parses CSV/XLSX uploads of Australian state-government skilled migration lists
(e.g. NSW Subclass 190, VIC ROL, QLD BSMQ).

Expected column patterns (flexible — autodetected):
    - anzsco_code | code | unit_group
    - title | occupation | role
    - status | demand | priority   (optional: open/closed/high_demand)
    - notes (optional)

Output schema:
    {
      "state": "NSW",
      "list_type": "190",                # 190 | 491 | dama | rol | bsmq | ...
      "as_of_date": "2026-06-01",
      "source_url": "...",
      "codes": [
          {"anzsco_code": "261313", "title": "Software Engineer",
           "status": "open", "demand": "high", "notes": ""},
          ...
      ],
    }
"""
from __future__ import annotations

import csv
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SOURCE_NAME = "State Nomination Lists (Admin Upload)"
DATA_QUALITY = "official_state_govt_data"

_STATE_ALIASES = {
    "new south wales": "NSW", "nsw": "NSW",
    "victoria": "VIC", "vic": "VIC",
    "queensland": "QLD", "qld": "QLD",
    "south australia": "SA", "sa": "SA",
    "western australia": "WA", "wa": "WA",
    "tasmania": "TAS", "tas": "TAS",
    "australian capital territory": "ACT", "act": "ACT",
    "northern territory": "NT", "nt": "NT",
}

_STATUS_ALIASES = {
    "open": "open", "available": "open", "accepting": "open",
    "closed": "closed", "paused": "closed", "not available": "closed",
    "high": "high_demand", "high demand": "high_demand", "priority": "high_demand",
    "limited": "limited", "low": "low_demand",
}


def _norm_state(s: str) -> Optional[str]:
    if not s:
        return None
    key = str(s).strip().lower()
    return _STATE_ALIASES.get(key)


def _norm_status(s: str) -> str:
    if not s:
        return "open"  # default — admin should mark closed explicitly
    key = str(s).strip().lower()
    return _STATUS_ALIASES.get(key, key.replace(" ", "_"))


def _norm_code(c: Any) -> Optional[str]:
    """Normalise ANZSCO code (6 digits, leading-zero preserved)."""
    if c is None:
        return None
    s = str(c).strip()
    if "." in s:
        s = s.split(".")[0]
    digits = re.sub(r"\D", "", s)
    if len(digits) == 6:
        return digits
    if len(digits) == 5:
        return "0" + digits  # leading zero often stripped by Excel
    if len(digits) == 4:
        return digits  # 4-digit unit group — let importer expand
    return None


def _detect_columns(header: List[str]) -> Dict[str, int]:
    """Map known field names to column indexes."""
    cols = {}
    for i, raw in enumerate(header):
        h = str(raw or "").strip().lower()
        if not h:
            continue
        if "anzsco" in h or h in ("code", "unit_group", "occupation_code"):
            cols.setdefault("code", i)
        elif h in ("title", "occupation", "role", "occupation_title", "occupation_name"):
            cols.setdefault("title", i)
        elif h in ("status", "demand", "priority"):
            cols.setdefault("status", i)
        elif "note" in h or h in ("comment", "remarks"):
            cols.setdefault("notes", i)
    return cols


def parse_csv(path: str, state: str, list_type: str = "190",
              source_url: str = "", as_of_date: Optional[str] = None) -> Dict[str, Any]:
    """Parse a CSV file of state nomination occupations."""
    p = Path(path)
    rows = []
    with p.open(encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, [])
        cols = _detect_columns(header)
        if "code" not in cols:
            raise ValueError(
                f"Missing ANZSCO code column in {p.name}. Found headers: {header}",
            )
        for row in reader:
            if not row or not row[cols["code"]]:
                continue
            code = _norm_code(row[cols["code"]])
            if not code:
                continue
            rows.append({
                "anzsco_code": code,
                "title": (row[cols["title"]] if "title" in cols and cols["title"] < len(row) else "").strip(),
                "status": _norm_status(row[cols["status"]] if "status" in cols and cols["status"] < len(row) else ""),
                "notes": (row[cols["notes"]] if "notes" in cols and cols["notes"] < len(row) else "").strip(),
            })
    return _build_envelope(state, list_type, source_url, as_of_date, rows)


def parse_xlsx(path: str, state: str, list_type: str = "190",
               source_url: str = "", as_of_date: Optional[str] = None) -> Dict[str, Any]:
    """Parse an XLSX file of state nomination occupations (first sheet)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, [])
    cols = _detect_columns(list(header))
    if "code" not in cols:
        wb.close()
        raise ValueError(
            f"Missing ANZSCO code column in {path}. Found headers: {list(header)}",
        )
    rows = []
    for raw in rows_iter:
        if not raw or len(raw) <= cols["code"] or raw[cols["code"]] in (None, ""):
            continue
        code = _norm_code(raw[cols["code"]])
        if not code:
            continue
        rows.append({
            "anzsco_code": code,
            "title": str(raw[cols["title"]] or "").strip() if "title" in cols and cols["title"] < len(raw) else "",
            "status": _norm_status(str(raw[cols["status"]] or "")) if "status" in cols and cols["status"] < len(raw) else "open",
            "notes": str(raw[cols["notes"]] or "").strip() if "notes" in cols and cols["notes"] < len(raw) else "",
        })
    wb.close()
    return _build_envelope(state, list_type, source_url, as_of_date, rows)


def _build_envelope(state: str, list_type: str, source_url: str,
                    as_of_date: Optional[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    norm_state = _norm_state(state)
    if not norm_state:
        raise ValueError(f"Unknown state code: {state}")
    # Dedupe by anzsco_code
    by_code = {}
    for r in rows:
        by_code[r["anzsco_code"]] = r
    return {
        "state": norm_state,
        "list_type": list_type.lower(),
        "as_of_date": as_of_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "source_url": source_url,
        "source_name": SOURCE_NAME,
        "data_quality": DATA_QUALITY,
        "codes": list(by_code.values()),
        "code_count": len(by_code),
    }


def parse_summary(path: str, state: str = "NSW", list_type: str = "190") -> Dict[str, Any]:
    """Sample 5 rows for parse-preview (auto-detects CSV vs XLSX)."""
    p = Path(path)
    if p.suffix.lower() in (".csv", ".tsv"):
        try:
            env = parse_csv(str(p), state, list_type)
        except Exception as e:  # noqa: BLE001
            return {"row_count": 0, "sample": [], "error": str(e),
                    "source": SOURCE_NAME, "source_url": ""}
    else:
        try:
            env = parse_xlsx(str(p), state, list_type)
        except Exception as e:  # noqa: BLE001
            return {"row_count": 0, "sample": [], "error": str(e),
                    "source": SOURCE_NAME, "source_url": ""}
    return {
        "row_count": env["code_count"],
        "source": SOURCE_NAME,
        "source_url": env.get("source_url", ""),
        "sample": env["codes"][:5],
        "honest_note": (f"Parsed {env['code_count']} codes for {env['state']} list {env['list_type']}. "
                        "Codes will be linked to AU occupation_master by ANZSCO code. "
                        "Phase 19.6 import_batch will track all updates (revocable 24h)."),
    }
